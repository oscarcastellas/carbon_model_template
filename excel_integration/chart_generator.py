"""
Chart Generation for Excel Interactive Modules

This module generates charts and visualizations that can be embedded in Excel.
"""

import sys
import os
from pathlib import Path
import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')  # Non-interactive backend
import matplotlib.pyplot as plt

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import xlwings as xw
    HAS_XLWINGS = True
except ImportError:
    HAS_XLWINGS = False


def create_sensitivity_heatmap(sensitivity_table: pd.DataFrame, output_path: str = None) -> str:
    """
    Create heatmap chart for sensitivity analysis.
    
    Parameters:
    -----------
    sensitivity_table : pd.DataFrame
        Sensitivity table with IRR values
    output_path : str, optional
        Path to save chart. If None, saves to temp file.
        
    Returns:
    --------
    str
        Path to saved chart image
    """
    if output_path is None:
        output_path = 'temp_sensitivity_heatmap.png'
    
    fig, ax = plt.subplots(figsize=(10, 8))
    
    # Convert table to numeric for plotting
    plot_data = sensitivity_table.copy()
    for col in plot_data.columns:
        plot_data[col] = pd.to_numeric(plot_data[col], errors='coerce')
    
    # Create heatmap
    im = ax.imshow(plot_data.values, cmap='RdYlGn', aspect='auto')
    
    # Set ticks
    ax.set_xticks(np.arange(len(plot_data.columns)))
    ax.set_yticks(np.arange(len(plot_data.index)))
    ax.set_xticklabels(plot_data.columns)
    ax.set_yticklabels(plot_data.index)
    
    # Rotate labels
    plt.setp(ax.get_xticklabels(), rotation=45, ha="right")
    
    # Add colorbar
    cbar = plt.colorbar(im, ax=ax)
    cbar.set_label('IRR (%)', rotation=270, labelpad=20)
    
    # Add text annotations
    for i in range(len(plot_data.index)):
        for j in range(len(plot_data.columns)):
            value = plot_data.iloc[i, j]
            if pd.notna(value):
                text = ax.text(j, i, f'{value:.1%}',
                             ha="center", va="center", color="black", fontsize=8)
    
    ax.set_title('Sensitivity Analysis - IRR Heatmap', fontsize=14, fontweight='bold')
    ax.set_xlabel('Price Multiplier', fontsize=12)
    ax.set_ylabel('Credit Volume Multiplier', fontsize=12)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    return output_path


def create_monte_carlo_histogram(irr_series: np.ndarray, npv_series: np.ndarray, output_dir: str = None) -> dict:
    """
    Create histogram charts for Monte Carlo results.
    
    Parameters:
    -----------
    irr_series : np.ndarray
        Array of IRR values
    npv_series : np.ndarray
        Array of NPV values
    output_dir : str, optional
        Directory to save charts. If None, uses current directory.
        
    Returns:
    --------
    dict
        Dictionary with paths to chart files
    """
    if output_dir is None:
        output_dir = '.'
    os.makedirs(output_dir, exist_ok=True)
    
    charts = {}
    
    # Filter out NaN values
    valid_irrs = irr_series[~pd.isna(irr_series) & np.isfinite(irr_series)]
    valid_npvs = npv_series[~pd.isna(npv_series) & np.isfinite(npv_series)]
    
    # IRR Histogram
    if len(valid_irrs) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(valid_irrs * 100, bins=50, edgecolor='black', alpha=0.7)
        ax.axvline(np.mean(valid_irrs) * 100, color='red', linestyle='--', linewidth=2, label=f'Mean: {np.mean(valid_irrs):.2%}')
        ax.axvline(np.percentile(valid_irrs, 10) * 100, color='orange', linestyle='--', linewidth=2, label=f'P10: {np.percentile(valid_irrs, 10):.2%}')
        ax.axvline(np.percentile(valid_irrs, 90) * 100, color='green', linestyle='--', linewidth=2, label=f'P90: {np.percentile(valid_irrs, 90):.2%}')
        ax.set_xlabel('IRR (%)', fontsize=12)
        ax.set_ylabel('Frequency', fontsize=12)
        ax.set_title('Monte Carlo Simulation - IRR Distribution', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        irr_path = os.path.join(output_dir, 'temp_mc_irr_histogram.png')
        plt.savefig(irr_path, dpi=150, bbox_inches='tight')
        plt.close()
        charts['irr_histogram'] = irr_path
    
    # NPV Histogram
    if len(valid_npvs) > 0:
        fig, ax = plt.subplots(figsize=(10, 6))
        ax.hist(valid_npvs / 1e6, bins=50, edgecolor='black', alpha=0.7)
        ax.axvline(np.mean(valid_npvs) / 1e6, color='red', linestyle='--', linewidth=2, label=f'Mean: ${np.mean(valid_npvs)/1e6:.1f}M')
        ax.axvline(np.percentile(valid_npvs, 10) / 1e6, color='orange', linestyle='--', linewidth=2, label=f'P10: ${np.percentile(valid_npvs, 10)/1e6:.1f}M')
        ax.axvline(np.percentile(valid_npvs, 90) / 1e6, color='green', linestyle='--', linewidth=2, label=f'P90: ${np.percentile(valid_npvs, 90)/1e6:.1f}M')
        ax.set_xlabel('NPV ($M)', fontsize=12)
        ax.set_ylabel('Frequency', fontsize=12)
        ax.set_title('Monte Carlo Simulation - NPV Distribution', fontsize=14, fontweight='bold')
        ax.legend()
        ax.grid(True, alpha=0.3)
        plt.tight_layout()
        npv_path = os.path.join(output_dir, 'temp_mc_npv_histogram.png')
        plt.savefig(npv_path, dpi=150, bbox_inches='tight')
        plt.close()
        charts['npv_histogram'] = npv_path
    
    return charts


def add_chart_to_excel(chart_path: str, sheet_name: str, cell_ref: str = 'E1'):
    """
    Add chart image to Excel sheet using xlwings.
    
    Parameters:
    -----------
    chart_path : str
        Path to chart image file
    sheet_name : str
        Name of Excel sheet
    cell_ref : str
        Cell reference where to place chart (default: 'E1')
    """
    if not HAS_XLWINGS:
        print("Warning: xlwings not available. Chart cannot be added to Excel.")
        return
    
    try:
        wb = xw.Book.caller()
        ws = wb.sheets[sheet_name]
        
        # Add picture to Excel
        ws.pictures.add(
            chart_path,
            left=ws.range(cell_ref).left,
            top=ws.range(cell_ref).top,
            width=400,  # pixels
            height=300  # pixels
        )
    except Exception as e:
        print(f"Warning: Could not add chart to Excel: {e}")


def embed_chart_in_excel_openpyxl(
    chart_path: str,
    excel_file: str,
    sheet_name: str,
    cell_ref: str = 'E1',
    width: int = 400,
    height: int = 300
) -> bool:
    """
    Embed chart image into Excel file using openpyxl (works without xlwings).
    
    Parameters:
    -----------
    chart_path : str
        Path to chart image file
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of Excel sheet
    cell_ref : str
        Cell reference where to place chart (default: 'E1')
    width : int
        Chart width in pixels (default: 400)
    height : int
        Chart height in pixels (default: 300)
        
    Returns:
    --------
    bool
        True if successful, False otherwise
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.drawing.image import Image
        import os
        
        if not os.path.exists(chart_path):
            print(f"Warning: Chart file not found: {chart_path}")
            return False
        
        if not os.path.exists(excel_file):
            print(f"Warning: Excel file not found: {excel_file}")
            return False
        
        # Load workbook
        wb = load_workbook(excel_file)
        
        if sheet_name not in wb.sheetnames:
            print(f"Warning: Sheet '{sheet_name}' not found in Excel file")
            wb.close()
            return False
        
        ws = wb[sheet_name]
        
        # Create image object
        img = Image(chart_path)
        img.width = width
        img.height = height
        
        # Add image to worksheet
        ws.add_image(img, cell_ref)
        
        # Save workbook
        wb.save(excel_file)
        wb.close()
        
        return True
        
    except ImportError:
        print("Warning: openpyxl not available. Chart cannot be embedded.")
        return False
    except Exception as e:
        print(f"Warning: Could not embed chart: {e}")
        return False


def create_deal_valuation_chart(
    price_points: np.ndarray,
    irr_points: np.ndarray,
    target_irr: float = None,
    output_path: str = None
) -> str:
    """
    Create Purchase Price vs IRR chart for deal valuation.
    
    Parameters:
    -----------
    price_points : np.ndarray
        Array of purchase price points
    irr_points : np.ndarray
        Array of corresponding IRR values
    target_irr : float, optional
        Target IRR to highlight on chart
    output_path : str, optional
        Path to save chart. If None, saves to temp file.
        
    Returns:
    --------
    str
        Path to saved chart image
    """
    if output_path is None:
        output_path = 'temp_deal_valuation_chart.png'
    
    fig, ax = plt.subplots(figsize=(10, 6))
    
    # Plot line
    ax.plot(price_points / 1e6, irr_points * 100, color='#366092', linewidth=2.5, marker='o', markersize=6)
    ax.fill_between(price_points / 1e6, irr_points * 100, alpha=0.2, color='#366092')
    
    # Add target IRR line if provided
    if target_irr is not None:
        ax.axhline(y=target_irr * 100, color='#FFC000', linewidth=2, linestyle='--', 
                  label=f'Target IRR: {target_irr:.1%}')
        ax.legend()
    
    ax.set_xlabel('Purchase Price ($M)', fontsize=12, fontweight='bold')
    ax.set_ylabel('IRR (%)', fontsize=12, fontweight='bold')
    ax.set_title('Purchase Price vs IRR Relationship', fontsize=14, fontweight='bold', pad=20)
    ax.grid(True, alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    return output_path


def create_breakeven_chart(
    breakeven_price: float = None,
    breakeven_volume: float = None,
    breakeven_streaming: float = None,
    output_path: str = None
) -> str:
    """
    Create breakeven analysis chart showing all breakeven points.
    
    Parameters:
    -----------
    breakeven_price : float, optional
        Breakeven price
    breakeven_volume : float, optional
        Breakeven volume multiplier
    breakeven_streaming : float, optional
        Breakeven streaming percentage
    output_path : str, optional
        Path to save chart. If None, saves to temp file.
        
    Returns:
    --------
    str
        Path to saved chart image
    """
    if output_path is None:
        output_path = 'temp_breakeven_chart.png'
    
    fig, ax = plt.subplots(figsize=(10, 6))
    
    labels = []
    values = []
    colors_list = []
    
    if breakeven_price is not None:
        labels.append('Breakeven Price\n($/ton)')
        values.append(breakeven_price)
        colors_list.append('#366092')
    
    if breakeven_volume is not None:
        labels.append('Breakeven Volume\n(Multiplier)')
        values.append(breakeven_volume * 100)  # Convert to percentage for display
        colors_list.append('#70AD47')
    
    if breakeven_streaming is not None:
        labels.append('Breakeven Streaming\n(%)')
        values.append(breakeven_streaming * 100)
        colors_list.append('#FFC000')
    
    if not labels:
        ax.text(0.5, 0.5, 'No breakeven data available', ha='center', va='center', 
               fontsize=12, transform=ax.transAxes)
        ax.set_title('Breakeven Analysis', fontsize=14, fontweight='bold', pad=20)
        plt.tight_layout()
        plt.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close()
        return output_path
    
    # Create bar chart
    x_pos = np.arange(len(labels))
    bars = ax.bar(x_pos, values, color=colors_list, edgecolor='black', linewidth=1.5, width=0.6)
    
    # Add value labels on bars
    for bar, val in zip(bars, values):
        if 'Price' in labels[bars.index(bar)]:
            label_text = f'${val:.2f}'
        else:
            label_text = f'{val:.1f}%'
        ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + max(values) * 0.05,
               label_text, ha='center', va='bottom', fontweight='bold', fontsize=11)
    
    ax.set_xticks(x_pos)
    ax.set_xticklabels(labels)
    ax.set_ylabel('Value', fontweight='bold')
    ax.set_title('Breakeven Analysis - Key Metrics', fontsize=14, fontweight='bold', pad=20)
    ax.grid(axis='y', alpha=0.3)
    
    plt.tight_layout()
    plt.savefig(output_path, dpi=150, bbox_inches='tight')
    plt.close()
    
    return output_path


if __name__ == '__main__':
    print("Chart generation module for Excel interactive modules")
    print("Functions available:")
    print("  - create_sensitivity_heatmap()")
    print("  - create_monte_carlo_histogram()")
    print("  - create_deal_valuation_chart()")
    print("  - create_breakeven_chart()")
    print("  - add_chart_to_excel()")
    print("  - embed_chart_in_excel_openpyxl()")

