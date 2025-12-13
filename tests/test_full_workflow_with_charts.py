#!/usr/bin/env python3
"""
Test Full Workflow with Charts

Tests the complete workflow:
1. GUI generates Excel from template (with charts in presentation sheets)
2. User fills input cells in analysis sheets
3. Python scripts run analysis and generate charts
"""

import sys
import os
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

from data.multi_file_loader import MultiFileLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from analysis.monte_carlo import MonteCarloSimulator
from analysis.sensitivity import SensitivityAnalyzer
from risk.flagger import RiskFlagger
from risk.scorer import RiskScoreCalculator
from valuation.breakeven import BreakevenCalculator
from valuation.deal_valuation import DealValuationSolver
from export.excel import ExcelExporter
import pandas as pd
import numpy as np


def test_full_workflow():
    """Test the complete workflow with charts."""
    print("=" * 70)
    print("TESTING FULL WORKFLOW WITH CHARTS")
    print("=" * 70)
    print()
    
    # Step 1: Load test data
    print("Step 1: Loading test data...")
    data_file = project_root / "Analyst_Model_Test_OCC.xlsx"
    
    if not data_file.exists():
        print(f"ERROR: Test data file not found: {data_file}")
        return False
    
    loader = MultiFileLoader()
    data = loader.load_excel(str(data_file))
    
    if data is None or data.empty:
        print("ERROR: Could not load data")
        return False
    
    print(f"   ✓ Data loaded: {len(data)} rows")
    print()
    
    # Step 2: Extract assumptions
    print("Step 2: Extracting assumptions...")
    from data.loader import DataLoader
    data_loader = DataLoader()
    assumptions = data_loader.extract_assumptions(str(data_file))
    
    # Set defaults if missing
    assumptions.setdefault('wacc', 0.08)
    assumptions.setdefault('rubicon_investment_total', 20_000_000)
    assumptions.setdefault('investment_tenor', 5)
    assumptions.setdefault('streaming_percentage_initial', 0.48)
    assumptions.setdefault('price_growth_base', 0.03)
    assumptions.setdefault('price_growth_std_dev', 0.02)
    assumptions.setdefault('volume_multiplier_base', 1.0)
    assumptions.setdefault('volume_std_dev', 0.15)
    
    print(f"   ✓ Assumptions extracted")
    print(f"      WACC: {assumptions.get('wacc', 0):.2%}")
    print(f"      Investment: ${assumptions.get('rubicon_investment_total', 0):,.0f}")
    print(f"      Streaming: {assumptions.get('streaming_percentage_initial', 0):.2%}")
    print()
    
    # Step 3: Run DCF analysis
    print("Step 3: Running DCF analysis...")
    dcf_calc = DCFCalculator(
        wacc=assumptions['wacc'],
        rubicon_investment_total=assumptions['rubicon_investment_total'],
        investment_tenor=assumptions['investment_tenor']
    )
    irr_calc = IRRCalculator()
    
    dcf_results = dcf_calc.run_dcf(data, assumptions['streaming_percentage_initial'])
    
    # Convert to valuation schedule format
    # run_dcf returns a dict with 'cash_flows', 'present_values', 'irr', 'npv', etc.
    cash_flows = dcf_results.get('cash_flows', pd.Series())
    present_values = dcf_results.get('present_values', pd.Series())
    
    # Create valuation schedule DataFrame
    valuation_schedule = pd.DataFrame({
        'cash_flow': cash_flows,
        'present_value': present_values
    })
    
    actual_irr = dcf_results['irr']
    target_irr = 0.20
    
    print(f"   ✓ DCF complete")
    print(f"      NPV: ${valuation_schedule['present_value'].sum():,.2f}")
    print(f"      IRR: {actual_irr:.2%}")
    print()
    
    # Step 4: Run risk analysis
    print("Step 4: Running risk analysis...")
    risk_flagger = RiskFlagger()
    risk_scorer = RiskScoreCalculator()
    
    risk_flags = risk_flagger.flag_risks(
        irr=actual_irr,
        npv=dcf_results['npv'],
        payback_period=None,
        credit_volumes=data['carbon_credits_gross'] if 'carbon_credits_gross' in data.columns else None,
        project_costs=data['project_implementation_costs'] if 'project_implementation_costs' in data.columns else None
    )
    
    risk_score = risk_scorer.calculate_overall_risk_score(
        irr=actual_irr,
        npv=dcf_results['npv'],
        payback_period=None,
        credit_volumes=data['carbon_credits_gross'] if 'carbon_credits_gross' in data.columns else None,
        base_prices=data['base_carbon_price'] if 'base_carbon_price' in data.columns else None,
        project_costs=data['project_implementation_costs'] if 'project_implementation_costs' in data.columns else None,
        total_investment=assumptions['rubicon_investment_total']
    )
    
    print(f"   ✓ Risk analysis complete")
    print(f"      Risk Level: {risk_flags.get('risk_level', 'Unknown')}")
    print(f"      Risk Score: {risk_score.get('overall_risk_score', 0):.0f}/100")
    print()
    
    # Step 5: Export to Excel (with charts in presentation sheets)
    print("Step 5: Exporting to Excel with charts...")
    output_file = project_root / "test_full_workflow_output.xlsx"
    
    excel_exporter = ExcelExporter()
    excel_exporter.export_model_to_excel(
        filename=str(output_file),
        assumptions=assumptions,
        target_streaming_percentage=assumptions['streaming_percentage_initial'],
        target_irr=target_irr,
        actual_irr=actual_irr,
        valuation_schedule=valuation_schedule,
        payback_period=None,
        risk_flags=risk_flags,
        risk_score=risk_score,
        use_template=True
    )
    
    print(f"   ✓ Excel file created: {output_file}")
    print()
    
    # Step 6: Test analysis scripts (simulate user filling inputs and running scripts)
    print("Step 6: Testing analysis scripts...")
    
    # Test Deal Valuation
    print("   Testing Deal Valuation...")
    try:
        from scripts.run_deal_valuation_from_excel import run_back_solver_from_excel
        # The script will read from Excel and write results
        run_back_solver_from_excel(str(output_file))
        print("   ✓ Deal Valuation script completed")
    except Exception as e:
        print(f"   ⚠ Deal Valuation script error: {e}")
    
    # Test Sensitivity Analysis
    print("   Testing Sensitivity Analysis...")
    try:
        from scripts.run_sensitivity_from_excel import run_sensitivity_from_excel
        run_sensitivity_from_excel(str(output_file))
        print("   ✓ Sensitivity Analysis script completed")
    except Exception as e:
        print(f"   ⚠ Sensitivity Analysis script error: {e}")
    
    # Test Monte Carlo (with fewer simulations for speed)
    print("   Testing Monte Carlo (100 sims for speed)...")
    try:
        # First, fill in Monte Carlo inputs in Excel
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        if 'Monte Carlo Results' in wb.sheetnames:
            ws = wb['Monte Carlo Results']
            ws['B8'] = 100  # Number of simulations
            ws['B9'] = 'Yes'  # Use GBM
            ws['B14'] = 0.03  # GBM Drift
            ws['B15'] = 0.15  # GBM Volatility
            wb.save(output_file)
            wb.close()
        
        from scripts.run_monte_carlo_from_excel import run_monte_carlo_from_excel
        run_monte_carlo_from_excel(str(output_file))
        print("   ✓ Monte Carlo script completed")
    except Exception as e:
        print(f"   ⚠ Monte Carlo script error: {e}")
    
    # Test Breakeven
    print("   Testing Breakeven Analysis...")
    try:
        from scripts.run_breakeven_from_excel import run_breakeven_from_excel
        run_breakeven_from_excel(str(output_file))
        print("   ✓ Breakeven Analysis script completed")
    except Exception as e:
        print(f"   ⚠ Breakeven Analysis script error: {e}")
    
    print()
    print("=" * 70)
    print("TEST COMPLETE")
    print("=" * 70)
    print()
    print(f"Output file: {output_file}")
    print()
    print("Please open the Excel file and verify:")
    print("  1. Presentation sheets (Inputs, Valuation, Summary) have charts")
    print("  2. Analysis sheets (Deal Valuation, Monte Carlo, Sensitivity, Breakeven) have:")
    print("     - Empty input cells (ready for user input)")
    print("     - Results populated (if scripts ran successfully)")
    print("     - Charts embedded (if scripts ran successfully)")
    print()
    
    return True


if __name__ == '__main__':
    success = test_full_workflow()
    sys.exit(0 if success else 1)

