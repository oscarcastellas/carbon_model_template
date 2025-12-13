#!/usr/bin/env python3
"""
Test Monte Carlo Interactive Sheet Creation
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from export.monte_carlo_interactive import InteractiveMonteCarloSheet
import xlsxwriter

def test_create_interactive_sheet():
    """Test creating interactive Excel sheet."""
    print("=" * 70)
    print("TESTING INTERACTIVE MONTE CARLO SHEET CREATION")
    print("=" * 70)
    print()
    
    # Create workbook
    output_file = "test_monte_carlo_interactive.xlsx"
    workbook = xlsxwriter.Workbook(output_file)
    
    # Create interactive sheet
    print("1. Creating interactive sheet...")
    creator = InteractiveMonteCarloSheet(workbook)
    assumptions = {
        'wacc': 0.08,
        'rubicon_investment_total': 20_000_000,
        'investment_tenor': 5,
        'streaming_percentage_initial': 0.48
    }
    
    sheet = creator.create_interactive_sheet(
        base_assumptions=assumptions,
        sheet_name="Monte Carlo Interactive"
    )
    print("   ✓ Interactive sheet created")
    print()
    
    # Close workbook
    workbook.close()
    
    print(f"2. Excel file created: {output_file}")
    
    # Verify file
    if os.path.exists(output_file):
        file_size = os.path.getsize(output_file) / 1024
        print(f"   ✓ File size: {file_size:.2f} KB")
        
        # Try to verify sheet exists
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_file)
            sheet_names = wb.sheetnames
            print(f"   ✓ Sheets: {', '.join(sheet_names)}")
            if 'Monte Carlo Interactive' in sheet_names:
                print("   ✓ Interactive sheet found!")
            wb.close()
        except ImportError:
            print("   ℹ Install openpyxl to verify sheet names")
    
    print()
    print("=" * 70)
    print("TEST COMPLETE")
    print("=" * 70)
    print()
    print("Next steps:")
    print(f"  1. Open {output_file}")
    print("  2. Go to 'Monte Carlo Interactive' sheet")
    print("  3. Set your simulation parameters (or use defaults)")
    print("  4. Run: python3 scripts/run_monte_carlo_from_excel.py " + output_file)
    print("  5. Check results in the sheet")


if __name__ == '__main__':
    test_create_interactive_sheet()

