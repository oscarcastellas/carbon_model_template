#!/usr/bin/env python3
"""
Run Interactive Analysis from Excel

Reads parameters from the 'Interactive Analysis' sheet in Excel,
runs the analysis, and updates results back to Excel.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from analysis_config import AnalysisConfig
from export.excel import ExcelExporter
from export.chart_exporter import ChartExporter
from export.interactive_sheet import InteractiveSheetCreator
from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from analysis.monte_carlo import MonteCarloSimulator
from analysis.gbm_simulator import GBMPriceSimulator
from risk.flagger import RiskFlagger
from risk.scorer import RiskScoreCalculator
from valuation.breakeven import BreakevenCalculator
from core.payback import PaybackCalculator
from analysis.volatility_visualizer import VolatilityVisualizer
import numpy as np


def read_parameters_from_excel(excel_file: str, sheet_name: str = "Interactive Analysis") -> dict:
    """
    Read parameters from Interactive Analysis sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of the sheet
        
    Returns:
    --------
    dict
        Dictionary of parameters
    """
    try:
        wb = openpyxl.load_workbook(excel_file, data_only=True)
        ws = wb[sheet_name]
        
        params = {}
        
        # Read all cells and find parameter values
        # Look for cells in column B (index 2) that correspond to labels in column A
        for row in range(1, ws.max_row + 1):
            label_cell = ws.cell(row, 1)  # Column A
            value_cell = ws.cell(row, 2)   # Column B
            
            if label_cell.value:
                label = str(label_cell.value).strip()
                value = value_cell.value
                
                # Map labels to parameter names
                if 'WACC' in label:
                    params['wacc'] = float(value) if value else 0.08
                elif 'Investment Total' in label:
                    params['rubicon_investment_total'] = float(value) if value else 20000000
                elif 'Tenor' in label:
                    params['investment_tenor'] = int(value) if value else 5
                elif 'Streaming Percentage' in label:
                    params['streaming_percentage_initial'] = float(value) if value else 0.48
                elif 'Use GBM' in label:
                    params['use_gbm'] = str(value).lower() in ['yes', 'true', '1']
                elif 'GBM Drift' in label:
                    params['gbm_drift'] = float(value) if value else 0.03
                elif 'GBM Volatility' in label:
                    params['gbm_volatility'] = float(value) if value else 0.15
                elif 'Number of Simulations' in label:
                    params['simulations'] = int(value) if value else 5000
                elif 'Price Growth Base' in label:
                    params['price_growth_base'] = float(value) if value else 0.03
                elif 'Price Growth Std Dev' in label:
                    params['price_growth_std_dev'] = float(value) if value else 0.02
                elif 'Volume Multiplier Base' in label:
                    params['volume_multiplier_base'] = float(value) if value else 1.0
                elif 'Volume Std Dev' in label:
                    params['volume_std_dev'] = float(value) if value else 0.15
        
        wb.close()
        
        # Set defaults if not found
        defaults = {
            'wacc': 0.08,
            'rubicon_investment_total': 20000000,
            'investment_tenor': 5,
            'streaming_percentage_initial': 0.48,
            'use_gbm': True,
            'gbm_drift': 0.03,
            'gbm_volatility': 0.15,
            'simulations': 5000,
            'price_growth_base': 0.03,
            'price_growth_std_dev': 0.02,
            'volume_multiplier_base': 1.0,
            'volume_std_dev': 0.15
        }
        
        for key, default in defaults.items():
            if key not in params:
                params[key] = default
        
        return params
        
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        print("Using default parameters...")
        return {
            'wacc': 0.08,
            'rubicon_investment_total': 20000000,
            'investment_tenor': 5,
            'streaming_percentage_initial': 0.48,
            'use_gbm': True,
            'gbm_drift': 0.03,
            'gbm_volatility': 0.15,
            'simulations': 5000,
            'price_growth_base': 0.03,
            'price_growth_std_dev': 0.02,
            'volume_multiplier_base': 1.0,
            'volume_std_dev': 0.15
        }


def update_excel_results(excel_file: str, results: dict, sheet_name: str = "Interactive Analysis"):
    """
    Update results in Excel file.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    results : dict
        Results dictionary
    sheet_name : str
        Name of the sheet to update
    """
    try:
        wb = openpyxl.load_workbook(excel_file)
        ws = wb[sheet_name]
        
        # Find "Current Results" section and update
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row, 1)
            if cell.value and 'Current Results' in str(cell.value):
                # Update results in next rows
                result_row = row + 1
                for key, label in [
                    ('mc_mean_irr', 'Mean IRR'),
                    ('mc_p10_irr', 'P10 IRR'),
                    ('mc_p90_irr', 'P90 IRR'),
                    ('mc_mean_npv', 'Mean NPV'),
                    ('mc_p10_npv', 'P10 NPV'),
                    ('mc_p90_npv', 'P90 NPV')
                ]:
                    if ws.cell(result_row, 1).value and label in str(ws.cell(result_row, 1).value):
                        value = results.get(key, 0)
                        if not (pd.isna(value) or not np.isfinite(value)):
                            ws.cell(result_row, 2).value = value
                    result_row += 1
                break
        
        wb.save(excel_file)
        wb.close()
        print(f"   ✓ Updated results in Excel file")
        
    except Exception as e:
        print(f"   ⚠ Could not update Excel file: {e}")
        print("   Results are shown below - you can manually update Excel")


def main():
    """Run interactive analysis from Excel."""
    print()
    print("="*70)
    print("INTERACTIVE ANALYSIS - READ FROM EXCEL & UPDATE RESULTS")
    print("="*70)
    print()
    
    # Excel file to read from
    excel_file = "carbon_model_with_charts.xlsx"
    
    if not os.path.exists(excel_file):
        print(f"❌ Error: {excel_file} not found!")
        print("   Please run: python3 examples/generate_excel_with_charts.py")
        return
    
    print(f"📖 Reading parameters from: {excel_file}")
    print("   Sheet: Interactive Analysis")
    print()
    
    # Read parameters from Excel
    params = read_parameters_from_excel(excel_file)
    
    print("📋 Parameters read from Excel:")
    print(f"   WACC: {params['wacc']:.2%}")
    print(f"   Investment: ${params['rubicon_investment_total']:,.0f}")
    print(f"   Tenor: {params['investment_tenor']} years")
    print(f"   Streaming: {params['streaming_percentage_initial']:.2%}")
    print(f"   Use GBM: {params['use_gbm']}")
    if params['use_gbm']:
        print(f"   GBM Drift: {params['gbm_drift']:.2%}")
        print(f"   GBM Volatility: {params['gbm_volatility']:.2%}")
    print(f"   Simulations: {params['simulations']:,}")
    print()
    
    # Load data
    print("📊 Loading data...")
    loader = DataLoader()
    data_file = "Analyst_Model_Test_OCC.xlsx"
    if not os.path.exists(data_file):
        print(f"   ⚠ {data_file} not found, using default data path")
    data = loader.load_data(data_file)
    print(f"   ✓ Loaded {len(data)} years of data")
    print()
    
    # Run DCF
    print("💰 Running DCF analysis...")
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=params['wacc'],
        rubicon_investment_total=params['rubicon_investment_total'],
        investment_tenor=params['investment_tenor'],
        irr_calculator=irr_calc
    )
    
    dcf_results = dcf_calc.run_dcf(data, params['streaming_percentage_initial'])
    print(f"   ✓ NPV: ${dcf_results['npv']:,.2f}")
    print(f"   ✓ IRR: {dcf_results['irr']:.2%}")
    print()
    
    # Run Monte Carlo
    print("🎲 Running Monte Carlo simulation...")
    print(f"   Method: {'GBM' if params['use_gbm'] else 'Growth-Rate'}")
    print(f"   Simulations: {params['simulations']:,}")
    print("   (This may take 1-3 minutes...)")
    print()
    
    mc_sim = MonteCarloSimulator(dcf_calc, irr_calc)
    mc_results = mc_sim.run_monte_carlo(
        base_data=data,
        streaming_percentage=params['streaming_percentage_initial'],
        price_growth_base=params['price_growth_base'],
        price_growth_std_dev=params['price_growth_std_dev'],
        volume_multiplier_base=params['volume_multiplier_base'],
        volume_std_dev=params['volume_std_dev'],
        simulations=params['simulations'],
        random_seed=42,
        use_percentage_variation=False,
        use_gbm=params['use_gbm'],
        gbm_drift=params['gbm_drift'],
        gbm_volatility=params['gbm_volatility']
    )
    
    print()
    print("   ✓ Monte Carlo complete!")
    print(f"   Mean IRR: {mc_results['mc_mean_irr']:.2%}")
    print(f"   P10 IRR: {mc_results['mc_p10_irr']:.2%}")
    print(f"   P90 IRR: {mc_results['mc_p90_irr']:.2%}")
    print()
    
    # Calculate other metrics
    print("📈 Calculating risk metrics...")
    payback_calc = PaybackCalculator()
    payback = payback_calc.calculate_payback_period(dcf_results['cash_flows'])
    
    risk_flagger = RiskFlagger()
    risk_flags = risk_flagger.flag_risks(
        dcf_results['irr'],
        dcf_results['npv'],
        payback,
        credit_volumes=data['carbon_credits_gross'],
        project_costs=data['project_implementation_costs']
    )
    
    risk_scorer = RiskScoreCalculator()
    risk_score = risk_scorer.calculate_overall_risk_score(
        dcf_results['irr'],
        dcf_results['npv'],
        payback,
        credit_volumes=data['carbon_credits_gross'],
        base_prices=data['base_carbon_price'],
        project_costs=data['project_implementation_costs'],
        total_investment=params['rubicon_investment_total']
    )
    
    breakeven_calc = BreakevenCalculator(dcf_calc, irr_calc)
    breakeven = breakeven_calc.calculate_all_breakevens(
        data, params['streaming_percentage_initial'], 0.0
    )
    
    print(f"   ✓ Risk Level: {risk_flags['risk_level'].upper()}")
    print(f"   ✓ Risk Score: {risk_score['overall_risk_score']:.1f}/100")
    print()
    
    # Update Excel file
    print("💾 Updating Excel file with new results...")
    update_excel_results(excel_file, mc_results)
    
    # Regenerate full Excel with updated results
    print("📊 Regenerating Excel file with all updated results...")
    
    assumptions = {
        'wacc': params['wacc'],
        'rubicon_investment_total': params['rubicon_investment_total'],
        'investment_tenor': params['investment_tenor'],
        'streaming_percentage_initial': params['streaming_percentage_initial'],
        'price_growth_base': params['price_growth_base'],
        'price_growth_std_dev': params['price_growth_std_dev'],
        'volume_multiplier_base': params['volume_multiplier_base'],
        'volume_std_dev': params['volume_std_dev'],
        'use_gbm': params['use_gbm'],
        'gbm_drift': params['gbm_drift'],
        'gbm_volatility': params['gbm_volatility'],
        'simulations': params['simulations']
    }
    
    # Generate charts if GBM is used
    if params['use_gbm']:
        print("   Generating updated charts...")
        visualizer = VolatilityVisualizer(output_dir="volatility_charts")
        base_prices = data['base_carbon_price']
        
        gbm_sim = GBMPriceSimulator()
        gbm_paths = []
        for i in range(1000):
            path = gbm_sim.generate_gbm_path_from_base(
                base_prices=base_prices,
                drift=params['gbm_drift'],
                volatility=params['gbm_volatility'],
                random_seed=None
            )
            gbm_paths.append(path)
        
        saved_charts = visualizer.generate_full_report(
            base_prices=base_prices,
            gbm_paths=gbm_paths,
            monte_carlo_results=mc_results,
            output_prefix="carbon_price_volatility"
        )
        print(f"   ✓ Generated {len(saved_charts)} charts")
    
    # Create updated Excel
    excel_exporter = ExcelExporter()
    import xlsxwriter
    
    workbook = xlsxwriter.Workbook(excel_file, {'nan_inf_to_errors': True})
    formats = excel_exporter._create_formats(workbook)
    
    # Create all sheets
    inputs_sheet = workbook.add_worksheet('Inputs & Assumptions')
    excel_exporter._write_inputs_sheet(
        workbook, inputs_sheet, formats, assumptions,
        params['streaming_percentage_initial'], 0.20
    )
    
    valuation_sheet = workbook.add_worksheet('Valuation Schedule')
    excel_exporter._write_valuation_schedule_with_formulas(
        workbook, valuation_sheet, formats, dcf_results['results_df'], inputs_sheet
    )
    
    summary_sheet = workbook.add_worksheet('Summary & Results')
    excel_exporter._write_summary_results_sheet(
        workbook, summary_sheet, formats, dcf_results['results_df'],
        inputs_sheet, dcf_results['irr'], payback, mc_results,
        risk_flags, risk_score, breakeven
    )
    
    mc_sheet = workbook.add_worksheet('Monte Carlo Results')
    excel_exporter._write_monte_carlo_sheet(
        workbook, mc_sheet, formats, mc_results
    )
    
    # Interactive sheet with updated results
    interactive_creator = InteractiveSheetCreator(workbook)
    interactive_sheet = interactive_creator.create_interactive_analysis_sheet(
        base_assumptions=assumptions,
        monte_carlo_results=mc_results,
        sheet_name="Interactive Analysis"
    )
    
    # Charts sheet
    if params['use_gbm'] and 'saved_charts' in locals():
        chart_exporter = ChartExporter(workbook)
        charts_sheet = chart_exporter.create_charts_sheet(
            charts=saved_charts,
            sheet_name="Volatility Charts"
        )
    
    workbook.close()
    
    print()
    print("="*70)
    print("ANALYSIS COMPLETE!")
    print("="*70)
    print()
    print(f"📊 Updated Excel file: {excel_file}")
    print()
    print("Results Summary:")
    print(f"   Mean IRR: {mc_results['mc_mean_irr']:.2%}")
    print(f"   P10 IRR: {mc_results['mc_p10_irr']:.2%}")
    print(f"   P90 IRR: {mc_results['mc_p90_irr']:.2%}")
    print(f"   Mean NPV: ${mc_results['mc_mean_npv']:,.0f}")
    print()
    print("Next steps:")
    print("   1. Open the Excel file")
    print("   2. Check 'Summary & Results' sheet for full analysis")
    print("   3. Adjust parameters in 'Interactive Analysis' sheet")
    print("   4. Run this script again to update results")
    print()
    print("="*70)


if __name__ == "__main__":
    main()

