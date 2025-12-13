#!/usr/bin/env python3
"""
Run Monte Carlo Simulation from Excel

This script reads input values from an Excel file, runs Monte Carlo simulation,
and writes results back to the Excel file.

Usage:
    python3 scripts/run_monte_carlo_from_excel.py [excel_file_path]
"""

import sys
import os
from pathlib import Path
from typing import Dict, Optional
import pandas as pd
import numpy as np

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from analysis.monte_carlo import MonteCarloSimulator


def read_inputs_from_excel(excel_file: str, sheet_name: str = "Monte Carlo Results") -> Dict:
    """
    Read input values from Excel interactive sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of the interactive sheet
        
    Returns:
    --------
    dict
        Dictionary with input values
    """
    wb = load_workbook(excel_file, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Helper function to safely read cell value
    def read_cell(cell_ref, default, cell_type=float):
        value = ws[cell_ref].value
        if value is None or value == '':
            return default
        try:
            if cell_type == bool:
                return str(value).lower() in ['yes', 'true', '1', 'y']
            elif cell_type == int:
                return int(value)
            else:
                return float(value)
        except (ValueError, TypeError):
            return default
    
    # Read input cells
    inputs = {
        'simulations': read_cell('B8', 5000, int),
        'streaming_percentage': read_cell('B9', 0.48),
        'random_seed': read_cell('B10', None, int) if ws['B10'].value not in [None, ''] else None,
        'use_gbm': read_cell('B12', True, bool),
        'gbm_drift': read_cell('B14', 0.03),
        'gbm_volatility': read_cell('B15', 0.15),
        'price_growth_base': read_cell('B17', 0.03),
        'price_growth_std_dev': read_cell('B18', 0.02),
        'use_percentage_variation': read_cell('B19', False, bool),
        'volume_multiplier_base': read_cell('B21', 1.0),
        'volume_std_dev': read_cell('B22', 0.15)
    }
    
    # Validate inputs
    if inputs['simulations'] <= 0:
        inputs['simulations'] = 5000
    if inputs['streaming_percentage'] <= 0 or inputs['streaming_percentage'] > 1:
        inputs['streaming_percentage'] = 0.48
    if inputs['gbm_drift'] < 0:
        inputs['gbm_drift'] = 0.03
    if inputs['gbm_volatility'] < 0:
        inputs['gbm_volatility'] = 0.15
    if inputs['volume_multiplier_base'] <= 0:
        inputs['volume_multiplier_base'] = 1.0
    if inputs['volume_std_dev'] < 0:
        inputs['volume_std_dev'] = 0.15
    
    wb.close()
    return inputs


def write_results_to_excel(
    excel_file: str,
    results: Dict,
    sheet_name: str = "Monte Carlo Results"
) -> None:
    """
    Write Monte Carlo results to Excel sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    results : dict
        Results dictionary from Monte Carlo
    sheet_name : str
        Name of the interactive sheet
    """
    wb = load_workbook(excel_file)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Write IRR results
    ws['B27'] = results.get('mc_mean_irr', '')
    ws['B27'].number_format = '0.00%'
    ws['B28'] = results.get('mc_p10_irr', '')
    ws['B28'].number_format = '0.00%'
    ws['B29'] = results.get('mc_p50_irr', '')
    ws['B29'].number_format = '0.00%'
    ws['B30'] = results.get('mc_p90_irr', '')
    ws['B30'].number_format = '0.00%'
    ws['B31'] = results.get('mc_std_irr', '')
    ws['B31'].number_format = '0.00%'
    
    # Min/Max IRR
    irr_series = results.get('irr_series', [])
    if len(irr_series) > 0:
        valid_irrs = [x for x in irr_series if pd.notna(x) and np.isfinite(x)]
        if len(valid_irrs) > 0:
            ws['B32'] = float(np.min(valid_irrs))
            ws['B32'].number_format = '0.00%'
            ws['B33'] = float(np.max(valid_irrs))
            ws['B33'].number_format = '0.00%'
    
    # Write NPV results
    ws['B35'] = results.get('mc_mean_npv', '')
    ws['B35'].number_format = '$#,##0.00'
    ws['B36'] = results.get('mc_p10_npv', '')
    ws['B36'].number_format = '$#,##0.00'
    ws['B37'] = results.get('mc_p50_npv', '')
    ws['B37'].number_format = '$#,##0.00'
    ws['B38'] = results.get('mc_p90_npv', '')
    ws['B38'].number_format = '$#,##0.00'
    ws['B39'] = results.get('mc_std_npv', '')
    ws['B39'].number_format = '$#,##0.00'
    
    # Min/Max NPV
    npv_series = results.get('npv_series', [])
    if len(npv_series) > 0:
        valid_npvs = [x for x in npv_series if pd.notna(x) and np.isfinite(x)]
        if len(valid_npvs) > 0:
            ws['B40'] = float(np.min(valid_npvs))
            ws['B40'].number_format = '$#,##0.00'
            ws['B41'] = float(np.max(valid_npvs))
            ws['B41'].number_format = '$#,##0.00'
    
    # Write probabilities
    irr_series = results.get('irr_series', [])
    npv_series = results.get('npv_series', [])
    
    if len(irr_series) > 0:
        valid_irrs = [x for x in irr_series if pd.notna(x) and np.isfinite(x)]
        if len(valid_irrs) > 0:
            prob_irr_20 = sum(1 for x in valid_irrs if x > 0.20) / len(valid_irrs)
            prob_irr_15 = sum(1 for x in valid_irrs if x > 0.15) / len(valid_irrs)
            ws['B43'] = prob_irr_20
            ws['B43'].number_format = '0.00%'
            ws['B44'] = prob_irr_15
            ws['B44'].number_format = '0.00%'
    
    if len(npv_series) > 0:
        valid_npvs = [x for x in npv_series if pd.notna(x) and np.isfinite(x)]
        if len(valid_npvs) > 0:
            prob_npv_0 = sum(1 for x in valid_npvs if x > 0) / len(valid_npvs)
            prob_npv_10m = sum(1 for x in valid_npvs if x > 10_000_000) / len(valid_npvs)
            ws['B45'] = prob_npv_0
            ws['B45'].number_format = '0.00%'
            ws['B46'] = prob_npv_10m
            ws['B46'].number_format = '0.00%'
    
    # Write status
    ws['B48'] = 'Success - Monte Carlo Simulation Complete'
    
    # Generate and add histogram charts
    print("   Generating charts...")
    try:
        from excel_integration.chart_generator import create_monte_carlo_histogram, embed_chart_in_excel_openpyxl
        irr_series = np.array(results.get('irr_series', []))
        npv_series = np.array(results.get('npv_series', []))
        
        if len(irr_series) > 0 and len(npv_series) > 0:
            charts = create_monte_carlo_histogram(irr_series, npv_series)
            
            # Embed charts using openpyxl
            if 'irr_histogram' in charts:
                embed_chart_in_excel_openpyxl(
                    charts['irr_histogram'], excel_file, sheet_name, 'E27', width=500, height=350
                )
                print(f"   ✓ IRR histogram embedded")
            
            if 'npv_histogram' in charts:
                embed_chart_in_excel_openpyxl(
                    charts['npv_histogram'], excel_file, sheet_name, 'E35', width=500, height=350
                )
                print(f"   ✓ NPV histogram embedded")
        else:
            print(f"   ⚠ No simulation data - skipping charts")
    except Exception as e:
        print(f"   ⚠ Could not generate charts: {e}")
        print(f"   (Results are still written to Excel)")
    
    wb.save(excel_file)
    wb.close()


def run_monte_carlo_from_excel(excel_file: str) -> None:
    """
    Main function to run Monte Carlo simulation from Excel inputs.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file with interactive sheet
    """
    print("=" * 70)
    print("MONTE CARLO SIMULATION - EXCEL INTEGRATION")
    print("=" * 70)
    print()
    
    # Step 1: Read inputs from Excel
    print("1. Reading inputs from Excel...")
    try:
        inputs = read_inputs_from_excel(excel_file)
        print(f"   ✓ Simulations: {inputs['simulations']:,}")
        print(f"   ✓ Streaming %: {inputs['streaming_percentage']:.2%}")
        print(f"   ✓ Use GBM: {inputs['use_gbm']}")
        if inputs['use_gbm']:
            print(f"   ✓ GBM Drift: {inputs['gbm_drift']:.2%}")
            print(f"   ✓ GBM Volatility: {inputs['gbm_volatility']:.2%}")
        else:
            print(f"   ✓ Price Growth Base: {inputs['price_growth_base']:.2%}")
            print(f"   ✓ Price Growth Std Dev: {inputs['price_growth_std_dev']:.2%}")
        print(f"   ✓ Volume Multiplier Base: {inputs['volume_multiplier_base']:.2f}")
        print(f"   ✓ Volume Std Dev: {inputs['volume_std_dev']:.2%}")
        if inputs['random_seed']:
            print(f"   ✓ Random Seed: {inputs['random_seed']}")
        print()
    except Exception as e:
        print(f"   ✗ Error reading inputs: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Step 2: Load data
    print("2. Loading project data...")
    data_file = None
    
    excel_dir = os.path.dirname(excel_file) or '.'
    possible_data_files = [
        os.path.join(excel_dir, "Analyst_Model_Test_OCC.xlsx"),
        "Analyst_Model_Test_OCC.xlsx",
        os.path.join(project_root, "Analyst_Model_Test_OCC.xlsx")
    ]
    
    for df in possible_data_files:
        if os.path.exists(df):
            data_file = df
            break
    
    if not data_file:
        print("   ✗ ERROR: Could not find data file (Analyst_Model_Test_OCC.xlsx)")
        wb = load_workbook(excel_file)
        ws = wb['Monte Carlo Results']
        ws['B48'] = 'Error - Data file not found'
        wb.save(excel_file)
        wb.close()
        return
    
    loader = DataLoader()
    data = loader.load_data(data_file)
    print(f"   ✓ Data loaded: {len(data)} years")
    print()
    
    # Step 3: Initialize DCF calculator
    print("3. Initializing DCF calculator...")
    wacc = 0.08
    investment_total = 20_000_000
    
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=wacc,
        rubicon_investment_total=investment_total,
        investment_tenor=5,
        irr_calculator=irr_calc
    )
    print("   ✓ DCF calculator initialized")
    print()
    
    # Step 4: Run Monte Carlo simulation
    print("4. Running Monte Carlo simulation...")
    print(f"   This will run {inputs['simulations']:,} simulations...")
    print(f"   This may take several minutes. Please wait...")
    print()
    
    try:
        # Set random seed if provided
        if inputs['random_seed']:
            np.random.seed(inputs['random_seed'])
        
        simulator = MonteCarloSimulator(dcf_calc, irr_calc)
        
        results = simulator.run_monte_carlo(
            base_data=data,
            streaming_percentage=inputs['streaming_percentage'],
            price_growth_base=inputs['price_growth_base'],
            price_growth_std_dev=inputs['price_growth_std_dev'],
            volume_multiplier_base=inputs['volume_multiplier_base'],
            volume_std_dev=inputs['volume_std_dev'],
            simulations=inputs['simulations'],
            random_seed=inputs['random_seed'],
            use_percentage_variation=inputs['use_percentage_variation'],
            use_gbm=inputs['use_gbm'],
            gbm_drift=inputs['gbm_drift'] if inputs['use_gbm'] else None,
            gbm_volatility=inputs['gbm_volatility'] if inputs['use_gbm'] else None
        )
        
        print(f"   ✓ Simulation complete!")
        print(f"   ✓ Mean IRR: {results['mc_mean_irr']:.2%}")
        print(f"   ✓ Mean NPV: ${results['mc_mean_npv']:,.2f}")
        print()
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        wb = load_workbook(excel_file)
        ws = wb['Monte Carlo Results']
        ws['B48'] = f'Error - {str(e)[:50]}'
        wb.save(excel_file)
        wb.close()
        return
    
    # Step 5: Write results to Excel
    print("5. Writing results to Excel...")
    try:
        write_results_to_excel(excel_file, results)
        print(f"   ✓ Results written to: {excel_file}")
        print()
    except Exception as e:
        print(f"   ✗ Error writing results: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("=" * 70)
    print("MONTE CARLO SIMULATION COMPLETE")
    print("=" * 70)
    print()
    print("Results have been written to the Excel file.")
    print("Open the file and check the 'Monte Carlo Results' sheet.")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = input("Enter path to Excel file (or press Enter for default): ").strip()
        if not excel_file:
            possible_files = [
                "test_monte_carlo_output.xlsx",
                "test_sensitivity_interactive.xlsx",
                "Full_Model_Export_With_Productivity_Tools.xlsx"
            ]
            for f in possible_files:
                if os.path.exists(f):
                    excel_file = f
                    break
        
        if not excel_file:
            print("ERROR: No Excel file specified and no default file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        sys.exit(1)
    
    run_monte_carlo_from_excel(excel_file)

