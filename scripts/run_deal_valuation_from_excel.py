#!/usr/bin/env python3
"""
Run Deal Valuation Back-Solver from Excel

This script reads input values from an Excel file, runs the back-solver,
and writes results back to the Excel file.

Usage:
    python3 scripts/run_deal_valuation_from_excel.py [excel_file_path]

Or from Excel:
    1. Open the Excel file with the interactive sheet
    2. Set your input variables
    3. Run this script (or use xlwings button)
    4. Results will be written back to Excel
"""

import sys
import os
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

from typing import Dict
from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from valuation.deal_valuation import DealValuationSolver


def read_inputs_from_excel(excel_file: str, sheet_name: str = "Deal Valuation") -> Dict:
    """
    Read input values from Excel interactive sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of the interactive sheet
        
    Returns:
    --------
    dict
        Dictionary with input values
    """
    wb = load_workbook(excel_file, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Helper function to safely read cell value
    def read_cell(cell_ref, default):
        value = ws[cell_ref].value
        if value is None or value == '':
            return default
        try:
            return type(default)(value)
        except (ValueError, TypeError):
            return default
    
    # Read input cells (based on named ranges or cell positions)
    inputs = {
        'target_irr': read_cell('B8', 0.20),  # Target IRR
        'streaming_percentage': read_cell('B9', 0.48),  # Streaming %
        'purchase_price': read_cell('B10', 20000000),  # Purchase Price
        'investment_tenor': read_cell('B11', 5),  # Investment Tenor
        'calc_type': str(read_cell('B13', 'Solve for Purchase Price')).strip()  # Calculation type
    }
    
    # Validate inputs
    if inputs['investment_tenor'] <= 0:
        inputs['investment_tenor'] = 5  # Default
    if inputs['target_irr'] <= 0:
        inputs['target_irr'] = 0.20  # Default
    if inputs['streaming_percentage'] <= 0:
        inputs['streaming_percentage'] = 0.48  # Default
    if inputs['purchase_price'] <= 0:
        inputs['purchase_price'] = 20000000  # Default
    
    # Normalize calculation type (handle variations)
    calc_type_lower = inputs['calc_type'].lower()
    if 'purchase price' in calc_type_lower or ('price' in calc_type_lower and 'solve' in calc_type_lower):
        inputs['calc_type'] = 'Solve for Purchase Price'
    elif 'irr' in calc_type_lower and 'calculate' in calc_type_lower:
        inputs['calc_type'] = 'Calculate IRR from Price'
    elif 'streaming' in calc_type_lower or 'streaming %' in calc_type_lower:
        inputs['calc_type'] = 'Solve for Streaming %'
    else:
        # Default to solving for purchase price
        inputs['calc_type'] = 'Solve for Purchase Price'
    
    wb.close()
    return inputs


def write_results_to_excel(
    excel_file: str,
    results: Dict,
    sheet_name: str = "Deal Valuation"
) -> None:
    """
    Write back-solver results to Excel sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    results : dict
        Results dictionary from back-solver
    sheet_name : str
        Name of the interactive sheet
    """
    wb = load_workbook(excel_file)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Unmerge result cells before writing (they might be merged from xlsxwriter)
    result_cells = ['B22', 'B23', 'B24', 'B25', 'B26', 'B27', 'B28', 'B30']
    for merged_range in list(ws.merged_cells.ranges):
        range_str = str(merged_range)
        # Check if any result cell is in this merged range
        if any(cell in range_str for cell in result_cells):
            try:
                ws.unmerge_cells(range_str)
            except:
                pass  # Ignore errors if already unmerged
    
    # Write results directly to cells
    # CORRECT CELL MAPPING:
    # B22 = Maximum Purchase Price
    # B23 = Actual IRR Achieved  
    # B24 = Target IRR
    # B25 = Difference
    # B26 = NPV at Calculated Price
    # B27 = Required Streaming %
    # B28 = Project IRR
    # B30 = Status
    
    try:
        print(f"   Writing results to Excel cells...")
        print(f"   Results keys: {list(results.keys())}")
        
        # Clear all result cells first
        for cell in result_cells:
            ws[cell] = ''
        
        if 'purchase_price' in results and 'target_irr' in results:
            # Solve for Purchase Price
            print(f"   Writing 'Solve for Purchase Price' results...")
            ws['B22'] = results.get('purchase_price', '')  # Maximum Purchase Price
            ws['B23'] = results.get('actual_irr', '')  # Actual IRR Achieved
            ws['B24'] = results.get('target_irr', '')  # Target IRR
            ws['B25'] = results.get('difference', '')  # Difference
            ws['B26'] = results.get('npv', '')  # NPV at Calculated Price
            ws['B27'] = ''  # Required Streaming % (not applicable)
            ws['B28'] = ''  # Project IRR (not applicable)
            ws['B30'] = 'Success - Purchase Price Calculated'
            print(f"   ✓ Written: B22=${results.get('purchase_price', 0):,.2f}, B23={results.get('actual_irr', 0):.2%}")
            
        elif 'irr' in results and 'purchase_price' not in results or ('irr' in results and 'purchase_price' in results and 'target_irr' not in results):
            # Calculate IRR from Price
            print(f"   Writing 'Calculate IRR from Price' results...")
            ws['B22'] = ''  # Maximum Purchase Price (not applicable)
            ws['B23'] = results.get('irr', '')  # Actual IRR Achieved (same as Project IRR)
            ws['B24'] = ''  # Target IRR (not applicable)
            ws['B25'] = ''  # Difference (not applicable)
            ws['B26'] = results.get('npv', '')  # NPV at Calculated Price
            ws['B27'] = ''  # Required Streaming % (not applicable)
            ws['B28'] = results.get('irr', '')  # Project IRR
            ws['B30'] = 'Success - IRR Calculated'
            print(f"   ✓ Written: B28={results.get('irr', 0):.2%}, B26=${results.get('npv', 0):,.2f}")
            
        elif 'streaming_percentage' in results:
            # Solve for Streaming %
            print(f"   Writing 'Solve for Streaming %' results...")
            ws['B22'] = ''  # Maximum Purchase Price (not applicable)
            ws['B23'] = results.get('actual_irr', '')  # Actual IRR Achieved
            ws['B24'] = results.get('target_irr', '')  # Target IRR
            ws['B25'] = results.get('difference', '')  # Difference
            ws['B26'] = results.get('npv', '')  # NPV at Calculated Price
            ws['B27'] = results.get('streaming_percentage', '')  # Required Streaming %
            ws['B28'] = ''  # Project IRR (not applicable)
            ws['B30'] = 'Success - Streaming % Calculated'
            print(f"   ✓ Written: B27={results.get('streaming_percentage', 0):.2%}, B23={results.get('actual_irr', 0):.2%}")
        else:
            print(f"   ⚠ Unknown result type, writing what we have...")
            # Write whatever we have
            if 'purchase_price' in results:
                ws['B22'] = results.get('purchase_price', '')
            if 'actual_irr' in results:
                ws['B23'] = results.get('actual_irr', '')
            if 'target_irr' in results:
                ws['B24'] = results.get('target_irr', '')
            if 'difference' in results:
                ws['B25'] = results.get('difference', '')
            if 'npv' in results:
                ws['B26'] = results.get('npv', '')
            if 'streaming_percentage' in results:
                ws['B27'] = results.get('streaming_percentage', '')
            if 'irr' in results:
                ws['B28'] = results.get('irr', '')
            ws['B30'] = 'Success - Results Written'
            
    except Exception as e:
        print(f"   ✗ Error writing to cells: {e}")
        import traceback
        traceback.print_exc()
        ws['B30'] = f'Error writing: {str(e)[:40]}'
    
    try:
        print(f"   Saving Excel file...")
        wb.save(excel_file)
        wb.close()
        print(f"   ✓ Excel file saved successfully")
    except Exception as e:
        print(f"   ✗ Error saving Excel file: {e}")
        import traceback
        traceback.print_exc()
        wb.close()
        raise


def run_back_solver_from_excel(excel_file: str) -> None:
    """
    Main function to run back-solver from Excel inputs.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file with interactive sheet
    """
    print("=" * 70)
    print("DEAL VALUATION BACK-SOLVER - EXCEL INTEGRATION")
    print("=" * 70)
    print()
    
    # Step 1: Read inputs from Excel
    print("1. Reading inputs from Excel...")
    try:
        inputs = read_inputs_from_excel(excel_file)
        print(f"   ✓ Target IRR: {inputs['target_irr']:.2%}")
        print(f"   ✓ Streaming %: {inputs['streaming_percentage']:.2%}")
        print(f"   ✓ Purchase Price: ${inputs['purchase_price']:,.0f}")
        print(f"   ✓ Investment Tenor: {inputs['investment_tenor']} years")
        print(f"   ✓ Calculation Type: {inputs['calc_type']}")
        print()
    except Exception as e:
        print(f"   ✗ Error reading inputs: {e}")
        return
    
    # Step 2: Load data (need to find data file or use existing data)
    print("2. Loading project data...")
    # Try to find data file - check if there's a data sheet or external file
    data_file = None
    
    # First, try to find Analyst_Model_Test_OCC.xlsx in the same directory
    excel_dir = os.path.dirname(excel_file) or '.'
    possible_data_files = [
        os.path.join(excel_dir, "Analyst_Model_Test_OCC.xlsx"),
        "Analyst_Model_Test_OCC.xlsx",
        os.path.join(project_root, "Analyst_Model_Test_OCC.xlsx")
    ]
    
    for df in possible_data_files:
        if os.path.exists(df):
            data_file = df
            break
    
    if not data_file:
        print("   ✗ ERROR: Could not find data file (Analyst_Model_Test_OCC.xlsx)")
        print("   Please ensure the data file is in the same directory as the Excel file")
        write_results_to_excel(excel_file, {}, sheet_name="Deal Valuation")
        wb = load_workbook(excel_file)
        ws = wb['Deal Valuation']
        ws['B30'] = 'Error - Data file not found'
        wb.save(excel_file)
        wb.close()
        return
    
    loader = DataLoader()
    data = loader.load_data(data_file)
    print(f"   ✓ Data loaded: {len(data)} years")
    print()
    
    # Step 3: Initialize DCF calculator
    print("3. Initializing DCF calculator...")
    # Get WACC and investment total from assumptions or use defaults
    wacc = 0.08  # Default, could be read from Excel
    investment_total = inputs.get('purchase_price', 20000000)  # Use purchase price or default
    
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=wacc,
        rubicon_investment_total=investment_total,
        investment_tenor=inputs['investment_tenor'],
        irr_calculator=irr_calc
    )
    print("   ✓ DCF calculator initialized")
    print()
    
    # Step 4: Run back-solver based on calculation type
    print(f"4. Running back-solver: {inputs['calc_type']}...")
    
    results = {}
    try:
        if inputs['calc_type'] == 'Solve for Purchase Price':
            # Create solver with temporary DCF (will be modified during solve)
            temp_dcf = DCFCalculator(
                wacc=wacc,
                rubicon_investment_total=20000000,  # Temporary
                investment_tenor=inputs['investment_tenor'],
                irr_calculator=IRRCalculator()
            )
            solver = DealValuationSolver(
                dcf_calculator=temp_dcf,
                data=data,
                tolerance=1e-4
            )
            results = solver.solve_for_purchase_price(
                target_irr=inputs['target_irr'],
                streaming_percentage=inputs['streaming_percentage'],
                investment_tenor=inputs['investment_tenor']
            )
            print(f"   ✓ Maximum Purchase Price: ${results['purchase_price']:,.2f}")
            print(f"   ✓ Actual IRR: {results['actual_irr']:.2%}")
            
        elif inputs['calc_type'] == 'Calculate IRR from Price':
            # Create solver with specified purchase price
            temp_dcf = DCFCalculator(
                wacc=wacc,
                rubicon_investment_total=inputs['purchase_price'],
                investment_tenor=inputs['investment_tenor'],
                irr_calculator=IRRCalculator()
            )
            solver = DealValuationSolver(
                dcf_calculator=temp_dcf,
                data=data,
                tolerance=1e-4
            )
            results = solver.solve_for_project_irr(
                purchase_price=inputs['purchase_price'],
                streaming_percentage=inputs['streaming_percentage'],
                investment_tenor=inputs['investment_tenor']
            )
            print(f"   ✓ Project IRR: {results['irr']:.2%}")
            print(f"   ✓ NPV: ${results['npv']:,.2f}")
            
        elif inputs['calc_type'] == 'Solve for Streaming %':
            # Create solver with specified purchase price
            temp_dcf = DCFCalculator(
                wacc=wacc,
                rubicon_investment_total=inputs['purchase_price'],
                investment_tenor=inputs['investment_tenor'],
                irr_calculator=IRRCalculator()
            )
            solver = DealValuationSolver(
                dcf_calculator=temp_dcf,
                data=data,
                tolerance=1e-4
            )
            results = solver.solve_for_streaming_given_price(
                purchase_price=inputs['purchase_price'],
                target_irr=inputs['target_irr'],
                investment_tenor=inputs['investment_tenor']
            )
            print(f"   ✓ Required Streaming %: {results['streaming_percentage']:.2%}")
            print(f"   ✓ Actual IRR: {results['actual_irr']:.2%}")
        else:
            raise ValueError(f"Unknown calculation type: {inputs['calc_type']}")
        
        print()
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        write_results_to_excel(excel_file, {}, sheet_name="Deal Valuation")
        wb = load_workbook(excel_file)
        ws = wb['Deal Valuation']
        ws['B30'] = f'Error - {str(e)[:50]}'
        wb.save(excel_file)
        wb.close()
        return
    
    # Step 5: Write results back to Excel
    print("5. Writing results to Excel...")
    try:
        write_results_to_excel(excel_file, results)
        print(f"   ✓ Results written to: {excel_file}")
        print()
    except Exception as e:
        print(f"   ✗ Error writing results: {e}")
        return
    
    # Step 6: Generate and embed charts
    print("6. Generating charts...")
    try:
        from excel_integration.chart_generator import create_deal_valuation_chart, embed_chart_in_excel_openpyxl
        import numpy as np
        
        # Create sample price points for chart (if we have purchase price data)
        if 'purchase_price' in results:
            # Generate chart showing price vs IRR relationship
            price_points = np.linspace(results.get('purchase_price', 0) * 0.5, 
                                     results.get('purchase_price', 0) * 1.5, 10)
            # For now, create a simple chart - in production, would calculate IRRs
            irr_points = np.array([results.get('actual_irr', 0.20)] * 10)
            
            chart_path = create_deal_valuation_chart(
                price_points, irr_points, 
                target_irr=inputs.get('target_irr', 0.20)
            )
            
            # Embed chart
            embed_chart_in_excel_openpyxl(
                chart_path, excel_file, "Deal Valuation", 'E15', width=500, height=350
            )
            print(f"   ✓ Chart embedded in Deal Valuation sheet")
        else:
            print(f"   ⚠ No purchase price data - skipping chart")
    except Exception as e:
        print(f"   ⚠ Could not generate chart: {e}")
        print(f"   (Results are still written to Excel)")
    
    print("=" * 70)
    print("BACK-SOLVER COMPLETE")
    print("=" * 70)
    print()
    print("Results have been written to the Excel file.")
    print("Open the file and check the 'Deal Valuation' sheet.")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        # Try to find the most recent Excel file or ask user
        excel_file = input("Enter path to Excel file (or press Enter for default): ").strip()
        if not excel_file:
            # Look for test files
            possible_files = [
                "test_back_solver_output.xlsx",
                "gui_test_back_solver_output.xlsx",
                "Full_Model_Export_With_Productivity_Tools.xlsx"
            ]
            for f in possible_files:
                if os.path.exists(f):
                    excel_file = f
                    break
        
        if not excel_file:
            print("ERROR: No Excel file specified and no default file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        sys.exit(1)
    
    run_back_solver_from_excel(excel_file)

