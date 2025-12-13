#!/usr/bin/env python3
"""
Run Breakeven Calculator from Excel

This script reads input values from an Excel file, runs breakeven analysis,
and writes results back to the Excel file.

Usage:
    python3 scripts/run_breakeven_from_excel.py [excel_file_path]
"""

import sys
import os
from pathlib import Path
from typing import Dict, Optional
import pandas as pd
import numpy as np

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import openpyxl
    from openpyxl import load_workbook
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from valuation.breakeven import BreakevenCalculator


def read_inputs_from_excel(excel_file: str, sheet_name: str = "Breakeven Analysis") -> Dict:
    """
    Read input values from Excel interactive sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of the interactive sheet
        
    Returns:
    --------
    dict
        Dictionary with input values
    """
    wb = load_workbook(excel_file, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Helper function to safely read cell value
    def read_cell(cell_ref, default, cell_type=str):
        value = ws[cell_ref].value
        if value is None or value == '':
            return default
        try:
            if cell_type == float:
                return float(value)
            elif cell_type == int:
                return int(value)
            else:
                return str(value).strip().lower()
        except (ValueError, TypeError):
            return default
    
    # Read input cells
    inputs = {
        'metric': read_cell('B8', 'all', str),
        'target_npv': read_cell('B9', 0.0, float),
        'streaming_percentage': read_cell('B10', 0.48, float)
    }
    
    # Normalize metric
    metric_lower = inputs['metric'].lower()
    if metric_lower in ['all', 'price', 'volume', 'streaming']:
        inputs['metric'] = metric_lower
    else:
        inputs['metric'] = 'all'
    
    # Validate inputs
    if inputs['streaming_percentage'] <= 0 or inputs['streaming_percentage'] > 1:
        inputs['streaming_percentage'] = 0.48
    
    wb.close()
    return inputs


def write_results_to_excel(
    excel_file: str,
    results: Dict,
    sheet_name: str = "Breakeven Analysis"
) -> None:
    """
    Write breakeven results to Excel sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    results : dict
        Results dictionary from breakeven calculator
    sheet_name : str
        Name of the interactive sheet
    """
    wb = load_workbook(excel_file)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Write breakeven price results
    if 'breakeven_price' in results:
        price_data = results['breakeven_price']
        if 'error' not in price_data:
            ws['B15'] = price_data.get('breakeven_price', '')
            ws['B15'].number_format = '$#,##0.00'
            ws['B16'] = price_data.get('base_price', '')
            ws['B16'].number_format = '$#,##0.00'
            ws['B17'] = price_data.get('price_multiplier', '')
            ws['B17'].number_format = '#,##0.00'
            ws['B18'] = price_data.get('target_npv', '')
            ws['B18'].number_format = '$#,##0.00'
        else:
            ws['B15'] = f"Error: {price_data.get('error', 'Unknown error')}"
    
    # Write breakeven volume results
    if 'breakeven_volume' in results:
        volume_data = results['breakeven_volume']
        if 'error' not in volume_data:
            ws['B20'] = volume_data.get('breakeven_volume_multiplier', '')
            ws['B20'].number_format = '#,##0.00'
            ws['B21'] = volume_data.get('base_volume', '')
            ws['B21'].number_format = '#,##0'
            if 'breakeven_volume' in volume_data:
                ws['B22'] = volume_data.get('breakeven_volume', '')
                ws['B22'].number_format = '#,##0'
            else:
                # Calculate breakeven volume from multiplier
                base_vol = volume_data.get('base_volume', 0)
                mult = volume_data.get('breakeven_volume_multiplier', 1.0)
                ws['B22'] = base_vol * mult
                ws['B22'].number_format = '#,##0'
            ws['B23'] = volume_data.get('target_npv', '')
            ws['B23'].number_format = '$#,##0.00'
        else:
            ws['B20'] = f"Error: {volume_data.get('error', 'Unknown error')}"
    
    # Write breakeven streaming results
    if 'breakeven_streaming' in results:
        streaming_data = results['breakeven_streaming']
        if 'error' not in streaming_data:
            ws['B25'] = streaming_data.get('breakeven_streaming', '')
            ws['B25'].number_format = '0.00%'
            # Get current streaming from inputs
            wb_inputs = load_workbook(excel_file, data_only=True)
            ws_inputs = wb_inputs[sheet_name]
            current_streaming = ws_inputs['B10'].value or 0.48
            ws['B26'] = current_streaming
            ws['B26'].number_format = '0.00%'
            wb_inputs.close()
            ws['B27'] = streaming_data.get('target_npv', '')
            ws['B27'].number_format = '$#,##0.00'
        else:
            ws['B25'] = f"Error: {streaming_data.get('error', 'Unknown error')}"
    
    # Write status
    ws['B29'] = 'Success - Breakeven Analysis Complete'
    
    # Generate and embed breakeven chart
    print("   Generating charts...")
    try:
        from excel_integration.chart_generator import create_breakeven_chart, embed_chart_in_excel_openpyxl
        
        # Extract breakeven values
        be_price = None
        be_volume = None
        be_streaming = None
        
        if 'breakeven_price' in results and results['breakeven_price']:
            be_price = results['breakeven_price'].get('breakeven_price')
        if 'breakeven_volume' in results and results['breakeven_volume']:
            be_volume = results['breakeven_volume'].get('breakeven_volume_multiplier')
        if 'breakeven_streaming' in results and results['breakeven_streaming']:
            be_streaming = results['breakeven_streaming'].get('breakeven_streaming')
        
        if be_price or be_volume or be_streaming:
            chart_path = create_breakeven_chart(be_price, be_volume, be_streaming)
            
            # Embed chart using openpyxl
            embed_chart_in_excel_openpyxl(
                chart_path, excel_file, sheet_name, 'E20', width=500, height=350
            )
            print(f"   ✓ Breakeven chart embedded")
        else:
            print(f"   ⚠ No breakeven data - skipping chart")
    except Exception as e:
        print(f"   ⚠ Could not generate chart: {e}")
        print(f"   (Results are still written to Excel)")
    
    wb.save(excel_file)
    wb.close()


def run_breakeven_from_excel(excel_file: str) -> None:
    """
    Main function to run breakeven analysis from Excel inputs.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file with interactive sheet
    """
    print("=" * 70)
    print("BREAKEVEN CALCULATOR - EXCEL INTEGRATION")
    print("=" * 70)
    print()
    
    # Step 1: Read inputs from Excel
    print("1. Reading inputs from Excel...")
    try:
        inputs = read_inputs_from_excel(excel_file)
        print(f"   ✓ Metric: {inputs['metric']}")
        print(f"   ✓ Target NPV: ${inputs['target_npv']:,.2f}")
        print(f"   ✓ Streaming %: {inputs['streaming_percentage']:.2%}")
        print()
    except Exception as e:
        print(f"   ✗ Error reading inputs: {e}")
        import traceback
        traceback.print_exc()
        return
    
    # Step 2: Load data
    print("2. Loading project data...")
    data_file = None
    
    excel_dir = os.path.dirname(excel_file) or '.'
    possible_data_files = [
        os.path.join(excel_dir, "Analyst_Model_Test_OCC.xlsx"),
        "Analyst_Model_Test_OCC.xlsx",
        os.path.join(project_root, "Analyst_Model_Test_OCC.xlsx")
    ]
    
    for df in possible_data_files:
        if os.path.exists(df):
            data_file = df
            break
    
    if not data_file:
        print("   ✗ ERROR: Could not find data file (Analyst_Model_Test_OCC.xlsx)")
        wb = load_workbook(excel_file)
        ws = wb['Breakeven Analysis']
        ws['B29'] = 'Error - Data file not found'
        wb.save(excel_file)
        wb.close()
        return
    
    loader = DataLoader()
    data = loader.load_data(data_file)
    print(f"   ✓ Data loaded: {len(data)} years")
    print()
    
    # Step 3: Initialize calculators
    print("3. Initializing calculators...")
    wacc = 0.08
    investment_total = 20_000_000
    
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=wacc,
        rubicon_investment_total=investment_total,
        investment_tenor=5,
        irr_calculator=irr_calc
    )
    breakeven_calc = BreakevenCalculator(dcf_calc, irr_calc)
    print("   ✓ Calculators initialized")
    print()
    
    # Step 4: Run breakeven analysis
    print(f"4. Running breakeven analysis ({inputs['metric']})...")
    
    try:
        results = {}
        
        if inputs['metric'] in ['all', 'price']:
            print("   Calculating breakeven price...")
            price_result = breakeven_calc.calculate_breakeven_price(
                data=data,
                streaming_percentage=inputs['streaming_percentage'],
                target_npv=inputs['target_npv']
            )
            results['breakeven_price'] = price_result
            if 'error' not in price_result:
                print(f"   ✓ Breakeven Price: ${price_result['breakeven_price']:,.2f}/ton")
                print(f"   ✓ Price Multiplier: {price_result['price_multiplier']:.2f}x")
        
        if inputs['metric'] in ['all', 'volume']:
            print("   Calculating breakeven volume...")
            volume_result = breakeven_calc.calculate_breakeven_volume(
                data=data,
                streaming_percentage=inputs['streaming_percentage'],
                target_npv=inputs['target_npv']
            )
            results['breakeven_volume'] = volume_result
            if 'error' not in volume_result:
                print(f"   ✓ Breakeven Volume Multiplier: {volume_result['breakeven_volume_multiplier']:.2f}x")
        
        if inputs['metric'] in ['all', 'streaming']:
            print("   Calculating breakeven streaming %...")
            streaming_result = breakeven_calc.calculate_breakeven_streaming(
                data=data,
                target_npv=inputs['target_npv']
            )
            results['breakeven_streaming'] = streaming_result
            if 'error' not in streaming_result:
                print(f"   ✓ Breakeven Streaming %: {streaming_result['breakeven_streaming']:.2%}")
        
        print()
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        wb = load_workbook(excel_file)
        ws = wb['Breakeven Analysis']
        ws['B29'] = f'Error - {str(e)[:50]}'
        wb.save(excel_file)
        wb.close()
        return
    
    # Step 5: Write results to Excel
    print("5. Writing results to Excel...")
    try:
        write_results_to_excel(excel_file, results)
        print(f"   ✓ Results written to: {excel_file}")
        print()
    except Exception as e:
        print(f"   ✗ Error writing results: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("=" * 70)
    print("BREAKEVEN ANALYSIS COMPLETE")
    print("=" * 70)
    print()
    print("Results have been written to the Excel file.")
    print("Open the file and check the 'Breakeven Analysis' sheet.")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = input("Enter path to Excel file (or press Enter for default): ").strip()
        if not excel_file:
            possible_files = [
                "test_breakeven_output.xlsx",
                "test_monte_carlo_interactive.xlsx",
                "Full_Model_Export_With_Productivity_Tools.xlsx"
            ]
            for f in possible_files:
                if os.path.exists(f):
                    excel_file = f
                    break
        
        if not excel_file:
            print("ERROR: No Excel file specified and no default file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        sys.exit(1)
    
    run_breakeven_from_excel(excel_file)

