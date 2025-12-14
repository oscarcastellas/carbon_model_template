#!/usr/bin/env python3
"""
Run Sensitivity Analysis from Excel

This script reads input values from an Excel file, runs sensitivity analysis,
and writes results back to the Excel file.

Usage:
    python3 scripts/run_sensitivity_from_excel.py [excel_file_path]
"""

import sys
import os
from pathlib import Path
from typing import Dict, List

# Add project root to path
project_root = Path(__file__).parent.parent
sys.path.insert(0, str(project_root))

try:
    import openpyxl
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
except ImportError:
    print("ERROR: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)

import pandas as pd
import numpy as np
from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from analysis.sensitivity import SensitivityAnalyzer


def read_inputs_from_excel(excel_file: str, sheet_name: str = "Sensitivity Analysis") -> Dict:
    """
    Read input values from Excel interactive sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sheet_name : str
        Name of the interactive sheet
        
    Returns:
    --------
    dict
        Dictionary with input values
    """
    wb = load_workbook(excel_file, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Helper function to safely read cell value
    def read_cell(cell_ref, default):
        value = ws[cell_ref].value
        if value is None or value == '':
            return default
        try:
            return type(default)(value)
        except (ValueError, TypeError):
            return default
    
    # Read input cells
    inputs = {
        'credit_min': read_cell('C8', 0.8),
        'credit_max': read_cell('E8', 1.2),
        'credit_step': read_cell('G8', 0.1),
        'price_min': read_cell('C9', 0.9),
        'price_max': read_cell('E9', 1.1),
        'price_step': read_cell('G9', 0.05),
        'streaming_percentage': read_cell('B10', 0.48)
    }
    
    # Validate inputs
    if inputs['credit_min'] >= inputs['credit_max']:
        raise ValueError("Credit min must be less than max")
    if inputs['price_min'] >= inputs['price_max']:
        raise ValueError("Price min must be less than max")
    if inputs['credit_step'] <= 0:
        inputs['credit_step'] = 0.1
    if inputs['price_step'] <= 0:
        inputs['price_step'] = 0.05
    if inputs['streaming_percentage'] <= 0 or inputs['streaming_percentage'] > 1:
        inputs['streaming_percentage'] = 0.48
    
    wb.close()
    return inputs


def generate_range(min_val: float, max_val: float, step: float) -> List[float]:
    """
    Generate range of values from min to max with given step.
    
    Parameters:
    -----------
    min_val : float
        Minimum value
    max_val : float
        Maximum value
    step : float
        Step size
        
    Returns:
    --------
    list
        List of values
    """
    values = []
    current = min_val
    while current <= max_val + step/2:  # Add small tolerance for floating point
        values.append(round(current, 6))
        current += step
    return values


def write_results_to_excel(
    excel_file: str,
    sensitivity_table: pd.DataFrame,
    summary_stats: Dict,
    sheet_name: str = "Sensitivity Analysis"
) -> None:
    """
    Write sensitivity analysis results to Excel sheet.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file
    sensitivity_table : pd.DataFrame
        Sensitivity table with IRR values
    summary_stats : dict
        Summary statistics
    sheet_name : str
        Name of the interactive sheet
    """
    
    wb = load_workbook(excel_file)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in Excel file")
    
    ws = wb[sheet_name]
    
    # Find where to write the table (after "Sensitivity Table" header)
    # Table starts at row 15 (0-indexed: 14)
    table_start_row = 14
    
    # Clear existing table area (rows 14-50, columns A-H)
    # Unmerge any merged cells first
    merged_ranges = list(ws.merged_cells.ranges)
    for merged_range in merged_ranges:
        if merged_range.min_row >= table_start_row + 1 and merged_range.max_row <= table_start_row + 50:
            ws.unmerge_cells(str(merged_range))
    
    for row in range(table_start_row, min(table_start_row + 50, ws.max_row + 1)):
        for col in range(1, 9):
            cell = ws.cell(row=row+1, column=col+1)
            if cell.value is not None:
                cell.value = None
            # Reset formatting
            cell.fill = None
            cell.font = None
    
    # Header fill and font
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=10)
    
    # Write table header
    # First cell: empty (top-left corner)
    header_cell = ws.cell(row=table_start_row+1, column=1)
    header_cell.value = 'Credit Volume →'
    header_cell.font = header_font
    header_cell.fill = header_fill
    header_cell.alignment = Alignment(horizontal='center', vertical='center')
    header_cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write column headers (price multipliers)
    col_idx = 2
    for price_mult in sensitivity_table.columns:
        cell = ws.cell(row=table_start_row+1, column=col_idx)
        cell.value = str(price_mult)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        col_idx += 1
    
    # Write row headers and data
    row_idx = table_start_row + 2
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for credit_mult in sensitivity_table.index:
        # Row header
        cell = ws.cell(row=row_idx, column=1)
        cell.value = str(credit_mult)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
        
        # Data cells
        col_idx = 2
        for price_mult in sensitivity_table.columns:
            irr_value = sensitivity_table.loc[credit_mult, price_mult]
            cell = ws.cell(row=row_idx, column=col_idx)
            if pd.notna(irr_value):
                cell.value = float(irr_value)
                cell.number_format = '0.00%'
            else:
                cell.value = 'N/A'
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            col_idx += 1
        row_idx += 1
    
    # Write summary statistics (check for merged cells first)
    summary_row = 30
    # Unmerge any cells in summary area
    for merged_range in list(ws.merged_cells.ranges):
        if merged_range.min_row >= summary_row and merged_range.max_row <= summary_row + 10:
            ws.unmerge_cells(str(merged_range))
    
    summary_cell = ws.cell(row=summary_row, column=2)
    if summary_cell.value is None or not hasattr(summary_cell, 'value') or not isinstance(summary_cell.value, str) or 'MergedCell' not in str(type(summary_cell)):
        summary_cell.value = summary_stats.get('min_irr', '')
        summary_cell.number_format = '0.00%'
    
    summary_cell = ws.cell(row=summary_row+1, column=2)
    if summary_cell.value is None or not hasattr(summary_cell, 'value') or not isinstance(summary_cell.value, str) or 'MergedCell' not in str(type(summary_cell)):
        summary_cell.value = summary_stats.get('max_irr', '')
        summary_cell.number_format = '0.00%'
    
    summary_cell = ws.cell(row=summary_row+2, column=2)
    if summary_cell.value is None or not hasattr(summary_cell, 'value') or not isinstance(summary_cell.value, str) or 'MergedCell' not in str(type(summary_cell)):
        summary_cell.value = summary_stats.get('irr_range', '')
        summary_cell.number_format = '0.00%'
    
    summary_cell = ws.cell(row=summary_row+3, column=2)
    if summary_cell.value is None or not hasattr(summary_cell, 'value') or not isinstance(summary_cell.value, str) or 'MergedCell' not in str(type(summary_cell)):
        summary_cell.value = summary_stats.get('base_case_irr', '')
        summary_cell.number_format = '0.00%'
    
    # Write status
    status_cell = ws.cell(row=35, column=2)
    if status_cell.value is None or not hasattr(status_cell, 'value') or not isinstance(status_cell.value, str) or 'MergedCell' not in str(type(status_cell)):
        status_cell.value = 'Success - Sensitivity Analysis Complete'
    
    # Generate and embed heatmap chart
    print("   Generating charts...")
    try:
        from excel_integration.chart_generator import create_sensitivity_heatmap, embed_chart_in_excel_openpyxl
        chart_path = create_sensitivity_heatmap(sensitivity_table)
        
        # Embed chart using openpyxl
        embed_chart_in_excel_openpyxl(
            chart_path, excel_file, sheet_name, 'E20', width=600, height=450
        )
        print(f"   ✓ Sensitivity heatmap embedded")
    except Exception as e:
        print(f"   ⚠ Could not generate chart: {e}")
        print(f"   (Results are still written to Excel)")
    
    wb.save(excel_file)
    wb.close()


def run_sensitivity_from_excel(excel_file: str) -> None:
    """
    Main function to run sensitivity analysis from Excel inputs.
    
    Parameters:
    -----------
    excel_file : str
        Path to Excel file with interactive sheet
    """
    
    print("=" * 70)
    print("SENSITIVITY ANALYSIS - EXCEL INTEGRATION")
    print("=" * 70)
    print()
    
    # Step 1: Read inputs from Excel
    print("1. Reading inputs from Excel...")
    try:
        inputs = read_inputs_from_excel(excel_file)
        print(f"   ✓ Credit Volume Range: {inputs['credit_min']:.2f}x to {inputs['credit_max']:.2f}x (step: {inputs['credit_step']:.2f})")
        print(f"   ✓ Price Multiplier Range: {inputs['price_min']:.2f}x to {inputs['price_max']:.2f}x (step: {inputs['price_step']:.2f})")
        print(f"   ✓ Streaming Percentage: {inputs['streaming_percentage']:.2%}")
        print()
    except Exception as e:
        print(f"   ✗ Error reading inputs: {e}")
        return
    
    # Step 2: Generate ranges
    print("2. Generating parameter ranges...")
    credit_range = generate_range(inputs['credit_min'], inputs['credit_max'], inputs['credit_step'])
    price_range = generate_range(inputs['price_min'], inputs['price_max'], inputs['price_step'])
    print(f"   ✓ Credit multipliers: {len(credit_range)} values")
    print(f"   ✓ Price multipliers: {len(price_range)} values")
    print(f"   ✓ Total scenarios: {len(credit_range) * len(price_range)}")
    print()
    
    # Step 3: Load data
    print("3. Loading project data...")
    data_file = None
    
    # Try to find data file
    excel_dir = os.path.dirname(excel_file) or '.'
    possible_data_files = [
        os.path.join(excel_dir, "Analyst_Model_Test_OCC.xlsx"),
        "Analyst_Model_Test_OCC.xlsx",
        os.path.join(project_root, "Analyst_Model_Test_OCC.xlsx")
    ]
    
    for df in possible_data_files:
        if os.path.exists(df):
            data_file = df
            break
    
    if not data_file:
        print("   ✗ ERROR: Could not find data file (Analyst_Model_Test_OCC.xlsx)")
        wb = load_workbook(excel_file)
        ws = wb['Sensitivity Analysis']
        ws['B35'] = 'Error - Data file not found'
        wb.save(excel_file)
        wb.close()
        return
    
    loader = DataLoader()
    data = loader.load_data(data_file)
    print(f"   ✓ Data loaded: {len(data)} years")
    print()
    
    # Step 4: Initialize DCF calculator
    print("4. Initializing DCF calculator...")
    wacc = 0.08  # Default, could be read from Excel
    investment_total = 20_000_000  # Default
    
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=wacc,
        rubicon_investment_total=investment_total,
        investment_tenor=5,  # Default
        irr_calculator=irr_calc
    )
    print("   ✓ DCF calculator initialized")
    print()
    
    # Step 5: Run sensitivity analysis
    print("5. Running sensitivity analysis...")
    print(f"   This may take a moment for {len(credit_range) * len(price_range)} scenarios...")
    
    try:
        sensitivity_analyzer = SensitivityAnalyzer(dcf_calc)
        sensitivity_table = sensitivity_analyzer.run_sensitivity_table(
            data=data,
            streaming_percentage=inputs['streaming_percentage'],
            credit_range=credit_range,
            price_range=price_range
        )
        
        print(f"   ✓ Sensitivity table generated: {sensitivity_table.shape[0]} rows × {sensitivity_table.shape[1]} columns")
        print()
        
        # Step 6: Calculate summary statistics
        print("6. Calculating summary statistics...")
        valid_irrs = sensitivity_table.values.flatten()
        valid_irrs = valid_irrs[~pd.isna(valid_irrs)]
        
        summary_stats = {
            'min_irr': float(np.min(valid_irrs)) if len(valid_irrs) > 0 else np.nan,
            'max_irr': float(np.max(valid_irrs)) if len(valid_irrs) > 0 else np.nan,
            'irr_range': float(np.max(valid_irrs) - np.min(valid_irrs)) if len(valid_irrs) > 0 else np.nan,
            'base_case_irr': float(sensitivity_table.loc['1.00x', '1.00x']) if '1.00x' in sensitivity_table.index and '1.00x' in sensitivity_table.columns else np.nan
        }
        
        print(f"   ✓ Min IRR: {summary_stats['min_irr']:.2%}")
        print(f"   ✓ Max IRR: {summary_stats['max_irr']:.2%}")
        print(f"   ✓ IRR Range: {summary_stats['irr_range']:.2%}")
        print(f"   ✓ Base Case IRR: {summary_stats['base_case_irr']:.2%}")
        print()
        
    except Exception as e:
        print(f"   ✗ Error: {e}")
        import traceback
        traceback.print_exc()
        wb = load_workbook(excel_file)
        ws = wb['Sensitivity Analysis']
        ws['B35'] = f'Error - {str(e)[:50]}'
        wb.save(excel_file)
        wb.close()
        return
    
    # Step 7: Write results to Excel
    print("7. Writing results to Excel...")
    try:
        write_results_to_excel(excel_file, sensitivity_table, summary_stats)
        print(f"   ✓ Results written to: {excel_file}")
        print()
    except Exception as e:
        print(f"   ✗ Error writing results: {e}")
        import traceback
        traceback.print_exc()
        return
    
    print("=" * 70)
    print("SENSITIVITY ANALYSIS COMPLETE")
    print("=" * 70)
    print()
    print("Results have been written to the Excel file.")
    print("Open the file and check the 'Sensitivity Analysis Interactive' sheet.")


if __name__ == '__main__':
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    else:
        excel_file = input("Enter path to Excel file (or press Enter for default): ").strip()
        if not excel_file:
            possible_files = [
                "test_sensitivity_output.xlsx",
                "test_interactive_deal_valuation.xlsx",
                "Full_Model_Export_With_Productivity_Tools.xlsx"
            ]
            for f in possible_files:
                if os.path.exists(f):
                    excel_file = f
                    break
        
        if not excel_file:
            print("ERROR: No Excel file specified and no default file found")
            sys.exit(1)
    
    if not os.path.exists(excel_file):
        print(f"ERROR: File not found: {excel_file}")
        sys.exit(1)
    
    run_sensitivity_from_excel(excel_file)

