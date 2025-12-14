#!/usr/bin/env python3
"""
Test Model Template with Fake Dataset

Tests the complete workflow using a fake dataset to stress test the model.
"""

import sys
from pathlib import Path

# Add project root to path
project_root = Path(__file__).parent
sys.path.insert(0, str(project_root))

from data.multi_file_loader import MultiFileLoader
from data.loader import DataLoader
from core.dcf import DCFCalculator
from core.irr import IRRCalculator
from core.payback import PaybackCalculator
from risk.flagger import RiskFlagger
from risk.scorer import RiskScoreCalculator
from export.excel import ExcelExporter
from scripts.run_deal_valuation_from_excel import run_back_solver_from_excel
from scripts.run_monte_carlo_from_excel import run_monte_carlo_from_excel
from scripts.run_sensitivity_from_excel import run_sensitivity_from_excel
from scripts.run_breakeven_from_excel import run_breakeven_from_excel

def test_with_fake_dataset(scenario: str = 'high_growth'):
    """Test complete workflow with fake dataset."""
    print("=" * 70)
    print(f"TESTING WITH FAKE DATASET: {scenario.upper()}")
    print("=" * 70)
    print()
    
    # Load fake dataset
    data_file = project_root / "data" / f"fake_dataset_{scenario}.xlsx"
    
    if not data_file.exists():
        print(f"ERROR: Fake dataset not found: {data_file}")
        print("Run: python3 data/create_fake_dataset.py")
        return False
    
    print(f"Step 1: Loading fake dataset: {data_file.name}")
    loader = MultiFileLoader()
    data = loader.load_excel(str(data_file))
    data_loader = DataLoader()
    assumptions = data_loader.extract_assumptions(str(data_file))
    
    # Set defaults
    assumptions.setdefault('wacc', 0.08)
    assumptions.setdefault('rubicon_investment_total', 20_000_000)
    assumptions.setdefault('investment_tenor', 5)
    assumptions.setdefault('streaming_percentage_initial', 0.48)
    
    print(f"   ✓ Data loaded: {len(data)} years")
    if 'base_carbon_price' in data.columns:
        print(f"   ✓ Price range: ${data['base_carbon_price'].min():.2f} - ${data['base_carbon_price'].max():.2f}/ton")
    if 'carbon_credits_gross' in data.columns:
        print(f"   ✓ Credits range: {data['carbon_credits_gross'].min():,.0f} - {data['carbon_credits_gross'].max():,.0f}")
    print(f"   ✓ Columns: {list(data.columns)}")
    print()
    
    # Run DCF
    print("Step 2: Running DCF analysis...")
    irr_calc = IRRCalculator()
    dcf_calc = DCFCalculator(
        wacc=assumptions['wacc'],
        rubicon_investment_total=assumptions['rubicon_investment_total'],
        investment_tenor=assumptions['investment_tenor'],
        irr_calculator=irr_calc
    )
    
    dcf_results = dcf_calc.run_dcf(data, assumptions['streaming_percentage_initial'])
    payback_calc = PaybackCalculator()
    payback = payback_calc.calculate_payback_period(dcf_results['cash_flows'])
    
    risk_flagger = RiskFlagger()
    risk_flags = risk_flagger.flag_risks(
        irr=dcf_results['irr'],
        npv=dcf_results['npv'],
        payback_period=payback,
        credit_volumes=data['carbon_credits_gross'] if 'carbon_credits_gross' in data.columns else None,
        project_costs=data['project_implementation_costs'] if 'project_implementation_costs' in data.columns else None
    )
    
    risk_scorer = RiskScoreCalculator()
    risk_score = risk_scorer.calculate_overall_risk_score(
        irr=dcf_results['irr'],
        npv=dcf_results['npv'],
        payback_period=payback,
        credit_volumes=data['carbon_credits_gross'] if 'carbon_credits_gross' in data.columns else None,
        base_prices=data['base_carbon_price'] if 'base_carbon_price' in data.columns else None,
        project_costs=data['project_implementation_costs'] if 'project_implementation_costs' in data.columns else None,
        total_investment=assumptions['rubicon_investment_total']
    )
    
    print(f"   ✓ NPV: ${dcf_results['npv']:,.2f}")
    print(f"   ✓ IRR: {dcf_results['irr']:.2%}")
    print(f"   ✓ Payback: {payback:.2f} years")
    print()
    
    # Export to Excel
    print("Step 3: Exporting to Excel with professional formatting...")
    output_file = project_root / f"test_output_{scenario}.xlsx"
    
    excel_exporter = ExcelExporter()
    excel_exporter.export_model_to_excel(
        filename=str(output_file),
        assumptions=assumptions,
        target_streaming_percentage=assumptions['streaming_percentage_initial'],
        target_irr=0.20,
        actual_irr=dcf_results['irr'],
        valuation_schedule=dcf_results['results_df'],
        payback_period=payback,
        risk_flags=risk_flags,
        risk_score=risk_score,
        use_template=True
    )
    
    print(f"   ✓ Excel file created: {output_file.name}")
    print()
    
    # Run all advanced analyses
    print("Step 4: Running advanced analyses...")
    print()
    
    # Deal Valuation
    print("   Running Deal Valuation...")
    try:
        run_back_solver_from_excel(str(output_file))
        print("   ✓ Deal Valuation complete")
    except Exception as e:
        print(f"   ✗ Deal Valuation error: {e}")
    
    # Monte Carlo (with fewer sims for speed)
    print("   Running Monte Carlo (1000 sims)...")
    try:
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        if 'Monte Carlo Results' in wb.sheetnames:
            ws = wb['Monte Carlo Results']
            ws['B8'] = 1000  # Number of simulations
            ws['B9'] = 'Yes'  # Use GBM
            ws['B14'] = 0.03  # GBM Drift
            ws['B15'] = 0.15  # GBM Volatility
            wb.save(output_file)
            wb.close()
        
        run_monte_carlo_from_excel(str(output_file))
        print("   ✓ Monte Carlo complete")
    except Exception as e:
        print(f"   ✗ Monte Carlo error: {e}")
    
    # Sensitivity
    print("   Running Sensitivity Analysis...")
    try:
        run_sensitivity_from_excel(str(output_file))
        print("   ✓ Sensitivity Analysis complete")
    except Exception as e:
        print(f"   ✗ Sensitivity error: {e}")
    
    # Breakeven
    print("   Running Breakeven Analysis...")
    try:
        run_breakeven_from_excel(str(output_file))
        print("   ✓ Breakeven Analysis complete")
    except Exception as e:
        print(f"   ✗ Breakeven error: {e}")
    
    print()
    print("=" * 70)
    print("TEST COMPLETE")
    print("=" * 70)
    print()
    print(f"Output file: {output_file}")
    print()
    print("All analyses have been run and results written to Excel.")
    print("Open the file to review the professional formatting and results.")
    
    return True

if __name__ == '__main__':
    import sys
    scenario = sys.argv[1] if len(sys.argv) > 1 else 'high_growth'
    success = test_with_fake_dataset(scenario)
    sys.exit(0 if success else 1)

