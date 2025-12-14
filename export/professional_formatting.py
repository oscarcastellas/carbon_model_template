"""
Professional Formatting Module

Provides comprehensive formatting functions to make Excel sheets look
professional and investor-ready with consistent styling, colors, and layouts.
"""

from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
    NamedStyle
)
from openpyxl.utils import get_column_letter
from typing import Optional


class ProfessionalFormatter:
    """Professional formatting styles for Excel sheets."""
    
    # Color palette (professional blue/gray theme)
    COLORS = {
        'header_blue': '366092',      # Dark blue for headers
        'header_light': '4472C4',      # Medium blue
        'accent_blue': 'D9E1F2',      # Light blue for labels
        'formula_green': 'E2EFDA',    # Light green for formulas
        'total_gray': 'F2F2F2',      # Light gray for totals
        'subtitle_gray': 'E7E6E6',    # Gray for subtitles
        'white': 'FFFFFF',
        'text_dark': '000000',
        'text_light': '666666',
        'success_green': 'C6EFCE',    # Green for positive
        'warning_yellow': 'FFEB9C',   # Yellow for warnings
        'error_red': 'FFC7CE',        # Red for errors
    }
    
    def __init__(self):
        """Initialize formatter with style definitions."""
        self.thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )
        
        self.medium_border = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
    
    def format_header_cell(self, cell, text: str = None):
        """Format a header cell with dark blue background and white text."""
        if text:
            cell.value = text
        cell.font = Font(bold=True, size=11, color=self.COLORS['white'])
        cell.fill = PatternFill(
            start_color=self.COLORS['header_blue'],
            end_color=self.COLORS['header_blue'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=True
        )
        cell.border = self.thin_border
        return cell
    
    def format_label_cell(self, cell, text: str = None):
        """Format a label cell with light blue background."""
        if text:
            cell.value = text
        cell.font = Font(bold=True, size=10, color=self.COLORS['text_dark'])
        cell.fill = PatternFill(
            start_color=self.COLORS['accent_blue'],
            end_color=self.COLORS['accent_blue'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        cell.border = self.thin_border
        return cell
    
    def format_formula_cell(self, cell, number_format: str = 'General'):
        """Format a formula cell with light green background."""
        cell.font = Font(size=10, color=self.COLORS['text_dark'])
        cell.fill = PatternFill(
            start_color=self.COLORS['formula_green'],
            end_color=self.COLORS['formula_green'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        cell.number_format = number_format
        cell.border = self.thin_border
        return cell
    
    def format_data_cell(self, cell, number_format: str = 'General'):
        """Format a data cell with white background."""
        cell.font = Font(size=10, color=self.COLORS['text_dark'])
        cell.fill = PatternFill(
            start_color=self.COLORS['white'],
            end_color=self.COLORS['white'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        cell.number_format = number_format
        cell.border = self.thin_border
        return cell
    
    def format_total_cell(self, cell, number_format: str = 'General'):
        """Format a total cell with gray background and bold text."""
        cell.font = Font(bold=True, size=10, color=self.COLORS['text_dark'])
        cell.fill = PatternFill(
            start_color=self.COLORS['total_gray'],
            end_color=self.COLORS['total_gray'],
            fill_type='solid'
        )
        cell.alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        cell.number_format = number_format
        cell.border = self.medium_border
        return cell
    
    def format_title_cell(self, cell, text: str):
        """Format a title cell."""
        cell.value = text
        cell.font = Font(bold=True, size=14, color=self.COLORS['text_dark'])
        cell.alignment = Alignment(horizontal='left', vertical='center')
        return cell
    
    def format_subtitle_cell(self, cell, text: str):
        """Format a subtitle cell."""
        cell.value = text
        cell.font = Font(bold=True, size=12, color=self.COLORS['text_dark'])
        cell.fill = PatternFill(
            start_color=self.COLORS['subtitle_gray'],
            end_color=self.COLORS['subtitle_gray'],
            fill_type='solid'
        )
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = self.thin_border
        return cell
    
    def format_valuation_schedule(self, ws):
        """Apply professional formatting to Valuation Schedule sheet."""
        # Title
        title_cell = ws.cell(1, 1)
        self.format_title_cell(title_cell, 'Valuation Schedule - 20 Year Cash Flow')
        
        # Header row (row 3)
        header_row = 3
        for col in range(2, 23):  # B to V (21 years)
            cell = ws.cell(header_row, col)
            year = 2025 + (col - 2)
            self.format_header_cell(cell, str(year))
        
        # Total column header
        total_col = 23  # Column W
        total_cell = ws.cell(header_row, total_col)
        self.format_header_cell(total_cell, 'Total')
        
        # Label column (column A, rows 4-14)
        label_fill = PatternFill(
            start_color=self.COLORS['accent_blue'],
            end_color=self.COLORS['accent_blue'],
            fill_type='solid'
        )
        for row in range(4, 15):
            cell = ws.cell(row, 1)
            if cell.value:
                cell.font = Font(bold=True, size=10)
                cell.fill = label_fill
                cell.alignment = Alignment(horizontal='right', vertical='center')
                cell.border = self.thin_border
        
        # Data cells (rows 4-14, columns B-V)
        formula_fill = PatternFill(
            start_color=self.COLORS['formula_green'],
            end_color=self.COLORS['formula_green'],
            fill_type='solid'
        )
        for row in range(4, 15):
            for col in range(2, 23):  # B to V
                cell = ws.cell(row, col)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    # Formula cell
                    cell.fill = formula_fill
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                elif cell.value is not None:
                    # Data cell
                    cell.border = self.thin_border
                    cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Total column (column W)
        total_fill = PatternFill(
            start_color=self.COLORS['total_gray'],
            end_color=self.COLORS['total_gray'],
            fill_type='solid'
        )
        for row in range(4, 15):
            cell = ws.cell(row, total_col)
            if cell.value:
                cell.font = Font(bold=True)
                cell.fill = total_fill
                cell.border = self.medium_border
                cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for col in range(2, 24):
            col_letter = get_column_letter(col)
            ws.column_dimensions[col_letter].width = 12
        ws.column_dimensions['W'].width = 15
    
    def format_summary_sheet(self, ws):
        """Apply professional formatting to Summary & Results sheet."""
        # Title
        title_cell = ws.cell(1, 1)
        if title_cell.value:
            self.format_title_cell(title_cell, title_cell.value)
        
        # Format all label cells (column A)
        label_fill = PatternFill(
            start_color=self.COLORS['accent_blue'],
            end_color=self.COLORS['accent_blue'],
            fill_type='solid'
        )
        for row in range(1, ws.max_row + 1):
            label_cell = ws.cell(row, 1)
            value_cell = ws.cell(row, 2)
            
            if label_cell.value and isinstance(label_cell.value, str):
                # Check if it's a subtitle
                if any(keyword in label_cell.value for keyword in ['Metrics', 'Assessment', 'Analysis', 'Summary']):
                    if 'Key' in label_cell.value or 'Risk' in label_cell.value or 'Monte' in label_cell.value or 'Breakeven' in label_cell.value:
                        self.format_subtitle_cell(label_cell, label_cell.value)
                else:
                    # Regular label
                    label_cell.font = Font(bold=True, size=10)
                    label_cell.fill = label_fill
                    label_cell.alignment = Alignment(horizontal='right', vertical='center')
                    label_cell.border = self.thin_border
                    
                    # Format corresponding value cell
                    if value_cell.value is not None:
                        value_cell.border = self.thin_border
                        value_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
    
    def format_analysis_sheet(self, ws, sheet_name: str):
        """Apply professional formatting to analysis sheets."""
        # Title
        title_cell = ws.cell(1, 1)
        if title_cell.value:
            self.format_title_cell(title_cell, title_cell.value)
        
        # Format input labels
        label_fill = PatternFill(
            start_color=self.COLORS['accent_blue'],
            end_color=self.COLORS['accent_blue'],
            fill_type='solid'
        )
        input_fill = PatternFill(
            start_color='FFF2CC',  # Light yellow for inputs
            end_color='FFF2CC',
            fill_type='solid'
        )
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, min(ws.max_column + 1, 10)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    # Check if it's an input label
                    if any(keyword in cell_str for keyword in ['target', 'streaming', 'purchase', 'simulation', 'gbm', 'volume', 'metric']):
                        if col == 1:  # Label column
                            cell.font = Font(bold=True, size=10)
                            cell.fill = label_fill
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.border = self.thin_border
                        elif col == 2:  # Input value column
                            cell.fill = input_fill
                            cell.border = self.thin_border
                            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Format result cells
        result_fill = PatternFill(
            start_color=self.COLORS['formula_green'],
            end_color=self.COLORS['formula_green'],
            fill_type='solid'
        )
        for row in range(1, ws.max_row + 1):
            for col in range(1, min(ws.max_column + 1, 10)):
                cell = ws.cell(row, col)
                if cell.value:
                    cell_str = str(cell.value).lower()
                    if any(keyword in cell_str for keyword in ['maximum', 'actual', 'difference', 'npv', 'mean', 'p10', 'p90', 'breakeven']):
                        if col == 1:  # Label
                            cell.font = Font(bold=True, size=10)
                            cell.fill = label_fill
                            cell.alignment = Alignment(horizontal='right', vertical='center')
                            cell.border = self.thin_border
                        elif col == 2:  # Value
                            cell.fill = result_fill
                            cell.border = self.thin_border
                            cell.alignment = Alignment(horizontal='right', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        if ws.max_column >= 3:
            ws.column_dimensions['C'].width = 30

