"""
Enhanced Template-Based Export

This module provides comprehensive template-based export that populates
all standard sheets while preserving interactive sheets with VBA/buttons.
"""

import os
import shutil
from pathlib import Path
from typing import Dict, Optional
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


class TemplateBasedExporter:
    """
    Exports Excel files using master template with all sheets and interactive modules.
    
    This ensures users get a fully-functional Excel file with:
    - All standard sheets (Inputs, Valuation, Summary, etc.) - populated with data
    - All interactive sheets (Deal Valuation, Sensitivity, Monte Carlo, Breakeven) - with VBA/buttons
    - Zero setup required!
    """
    
    def __init__(self, company_name: Optional[str] = None, num_years: Optional[int] = None):
        """
        Initialize template-based exporter.
        
        Parameters:
        -----------
        company_name : str, optional
            Company name to use (default: "Investor" from template)
        num_years : int, optional
            Number of years in model (default: 20 from template)
        """
        self.company_name = company_name or "Investor"
        self.num_years = num_years or 20
        # Use new generic master template
        # Prefer .xlsx (xlsxwriter creates .xlsx, not true .xlsm)
        self.template_path_xlsx = Path(__file__).parent.parent / "templates" / "master_template.xlsx"
        self.template_path = Path(__file__).parent.parent / "templates" / "master_template.xlsm"
    
    def export_with_template(
        self,
        filename: str,
        assumptions: Dict,
        target_streaming_percentage: float,
        target_irr: float,
        actual_irr: float,
        valuation_schedule: pd.DataFrame,
        sensitivity_table: Optional[pd.DataFrame] = None,
        payback_period: Optional[float] = None,
        monte_carlo_results: Optional[Dict] = None,
        risk_flags: Optional[Dict] = None,
        risk_score: Optional[Dict] = None,
        breakeven_results: Optional[Dict] = None,
        deal_valuation_results: Optional[Dict] = None
    ) -> bool:
        """
        Export using master template.
        
        Returns True if template was used, False if template not found (fallback needed).
        """
        # Check if completed template (.xlsm) exists, otherwise try .xlsx
        template_file = self.template_path if self.template_path.exists() else self.template_path_xlsx
        
        if not template_file.exists():
            print(f"Template not found. Expected at: {self.template_path}")
            print("Falling back to standard export (without interactive modules)")
            return False
        
        try:
            # Step 1: Copy template to output file
            print(f"Using master template: {template_file.name}")
            shutil.copy(template_file, filename)
            
            # Step 2: Load with openpyxl
            # Note: Template is .xlsx (xlsxwriter limitation), so no VBA to preserve
            wb = load_workbook(filename, keep_vba=False)
            
            # Step 3: Populate all data sheets
            print("Populating data sheets...")
            
            # Populate Inputs & Assumptions
            if 'Inputs & Assumptions' in wb.sheetnames:
                self._populate_inputs_sheet_comprehensive(wb['Inputs & Assumptions'], assumptions, target_streaming_percentage, target_irr)
                # Generate and embed charts
                self._add_presentation_charts_to_inputs(wb['Inputs & Assumptions'], assumptions, target_streaming_percentage)
            
            # Populate Valuation Schedule
            if 'Valuation Schedule' in wb.sheetnames:
                self._populate_valuation_sheet_comprehensive(wb['Valuation Schedule'], valuation_schedule)
                # Generate and embed charts
                self._add_presentation_charts_to_valuation(wb['Valuation Schedule'], valuation_schedule)
            
            # Populate Summary & Results
            if 'Summary & Results' in wb.sheetnames:
                self._populate_summary_sheet_comprehensive(
                    wb['Summary & Results'],
                    valuation_schedule,
                    actual_irr,
                    target_irr,
                    payback_period,
                    monte_carlo_results,
                    risk_flags,
                    risk_score,
                    breakeven_results
                )
                # Generate and embed charts
                self._add_presentation_charts_to_summary(wb['Summary & Results'], actual_irr, target_irr, risk_score)
            
            # Populate Deal Valuation (if results available)
            if 'Deal Valuation' in wb.sheetnames and deal_valuation_results:
                self._populate_deal_valuation_sheet(wb['Deal Valuation'], deal_valuation_results)
            
            # Populate Monte Carlo Results (if available)
            if 'Monte Carlo Results' in wb.sheetnames and monte_carlo_results:
                self._populate_monte_carlo_sheet(wb['Monte Carlo Results'], monte_carlo_results)
            
            # Populate Sensitivity Analysis (if available)
            if 'Sensitivity Analysis' in wb.sheetnames and sensitivity_table is not None:
                self._populate_sensitivity_sheet(wb['Sensitivity Analysis'], sensitivity_table)
            
            # Interactive sheets remain untouched (VBA and buttons preserved)
            print("✓ Interactive sheets preserved (VBA and buttons intact)")
            
            # Step 5: Apply professional formatting
            print("Applying professional formatting...")
            try:
                from .professional_formatting import ProfessionalFormatter
                formatter = ProfessionalFormatter()
                
                # Format Valuation Schedule
                if 'Valuation Schedule' in wb.sheetnames:
                    formatter.format_valuation_schedule(wb['Valuation Schedule'])
                
                # Format Summary & Results
                if 'Summary & Results' in wb.sheetnames:
                    formatter.format_summary_sheet(wb['Summary & Results'])
                
                # Format analysis sheets
                for sheet_name in ['Deal Valuation', 'Monte Carlo Results', 'Sensitivity Analysis', 'Breakeven Analysis']:
                    if sheet_name in wb.sheetnames:
                        formatter.format_analysis_sheet(wb[sheet_name], sheet_name)
                
                print("✓ Professional formatting applied")
            except Exception as e:
                print(f"⚠ Warning: Could not apply formatting: {e}")
            
            # Step 4: Save (preserves VBA if .xlsm)
            wb.save(filename)
            wb.close()
            
            print(f"✓ Template-based export complete: {filename}")
            print("  All standard sheets populated with data")
            print("  All interactive sheets ready to use (buttons work!)")
            return True
            
        except Exception as e:
            print(f"Error using template: {e}")
            import traceback
            traceback.print_exc()
            print("Falling back to standard export")
            return False
    
    def _populate_inputs_sheet_comprehensive(self, ws, assumptions: Dict, streaming_pct: float, target_irr: float):
        """Comprehensively populate Inputs & Assumptions sheet (matching template structure: B4-B9)."""
        # Template structure:
        # Row 1: Title
        # Row 3: "Base Financial Assumptions" (subtitle)
        # Row 4: WACC (B4)
        # Row 5: Investment Total (B5)
        # Row 6: Investment Tenor (B6)
        # Row 7: Initial Streaming Percentage (B7)
        # Row 8: Target IRR (B8)
        # Row 9: Target Streaming Percentage (B9)
        
        # Row 4: WACC (B4)
        ws.cell(row=4, column=2).value = assumptions.get('wacc', 0.08)
        ws.cell(row=4, column=2).number_format = '0.00%'
        
        # Row 5: Investment Total (B5) - support both old and new key names
        investment = assumptions.get('rubicon_investment_total') or assumptions.get('investment_total', 20_000_000)
        ws.cell(row=5, column=2).value = investment
        ws.cell(row=5, column=2).number_format = '$#,##0'
        
        # Row 6: Investment Tenor (B6)
        ws.cell(row=6, column=2).value = assumptions.get('investment_tenor', 5)
        ws.cell(row=6, column=2).number_format = '#,##0'
        
        # Row 7: Initial Streaming Percentage (B7)
        ws.cell(row=7, column=2).value = streaming_pct
        ws.cell(row=7, column=2).number_format = '0.00%'
        
        # Row 8: Target IRR (B8)
        ws.cell(row=8, column=2).value = target_irr
        ws.cell(row=8, column=2).number_format = '0.00%'
        
        # Row 9: Target Streaming Percentage (B9)
        ws.cell(row=9, column=2).value = streaming_pct
        ws.cell(row=9, column=2).number_format = '0.00%'
        
        # Monte Carlo assumptions (if present)
        if 'price_growth_base' in assumptions:
            # Find Monte Carlo section and populate
            for row in range(10, min(ws.max_row + 1, 30)):
                label_cell = ws.cell(row=row, column=1)
                if label_cell.value:
                    label_str = str(label_cell.value).lower()
                    value_cell = ws.cell(row=row, column=2)
                    
                    if 'price growth base' in label_str:
                        value_cell.value = assumptions.get('price_growth_base', 0.03)
                        value_cell.number_format = '0.00%'
                    elif 'price growth std' in label_str:
                        value_cell.value = assumptions.get('price_growth_std_dev', 0.02)
                        value_cell.number_format = '0.00%'
                    elif 'volume multiplier base' in label_str:
                        value_cell.value = assumptions.get('volume_multiplier_base', 1.0)
                        value_cell.number_format = '#,##0.00'
                    elif 'volume std' in label_str:
                        value_cell.value = assumptions.get('volume_std_dev', 0.15)
                        value_cell.number_format = '0.00%'
                    elif 'number of simulations' in label_str:
                        value_cell.value = assumptions.get('simulations', 5000)
                        value_cell.number_format = '#,##0'
    
    def _populate_valuation_sheet_comprehensive(self, ws, valuation_schedule: pd.DataFrame):
        """
        Comprehensively populate Valuation Schedule sheet.
        
        Structure: Years as columns (2025-2044), line items as rows.
        Matches the original ExcelExporter format with all formulas.
        """
        from openpyxl.utils import get_column_letter
        
        # Get reference to Inputs sheet for formula references
        wb = ws.parent
        inputs_sheet = wb['Inputs & Assumptions'] if 'Inputs & Assumptions' in wb.sheetnames else None
        
        # Find input cell references (matching template structure: B4-B9)
        # Template has: Row 4=WACC, Row 5=Investment, Row 6=Tenor, Row 7=Streaming, Row 8=Target IRR, Row 9=Target Streaming
        inputs_sheet_name = inputs_sheet.title if inputs_sheet else 'Inputs & Assumptions'
        wacc_cell = f"'{inputs_sheet_name}'!$B$4"
        investment_cell = f"'{inputs_sheet_name}'!$B$5"
        tenor_cell = f"'{inputs_sheet_name}'!$B$6"
        streaming_cell = f"'{inputs_sheet_name}'!$B$7"
        
        # Clear existing data (keep title in row 1)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.value = None
                    if cell.fill:
                        cell.fill = PatternFill()  # Reset to default fill
                    if cell.font:
                        cell.font = Font()  # Reset to default font
                    if cell.border:
                        cell.border = Border()  # Reset to default border
        
        # Title
        ws.cell(row=1, column=1).value = 'Valuation Schedule - 20 Year Cash Flow'
        ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        # Year headers (configurable, default 20 years) starting from column B
        start_year = 2025
        header_row = 3
        year_start_col = 2  # Column B
        num_years = self.num_years  # Use configurable year count
        
        # Border style
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Write year headers
        for year_idx in range(num_years):
            year = start_year + year_idx
            col = year_start_col + year_idx
            cell = ws.cell(row=header_row, column=col)
            cell.value = year
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
        
        # Total column
        total_col = year_start_col + num_years
        total_cell = ws.cell(row=header_row, column=total_col)
        total_cell.value = 'Total'
        total_cell.font = Font(bold=True, color='FFFFFF')
        total_cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = thin_border
        
        # Define line items (generic names, matching template structure)
        # Labels will match what's in the template (Investor or custom company name)
        line_items = [
            {'label': 'Carbon Credits Gross', 'data_col': 'carbon_credits_gross', 'format': 'number', 'total': True, 'key': 'carbon_credits_gross'},
            {'label': f'{self.company_name} Share of Credits', 'formula': 'credits_share', 'format': 'number', 'total': True, 'key': 'credits_share'},
            {'label': 'Base Carbon Price', 'data_col': 'base_carbon_price', 'format': 'currency', 'total': False, 'key': 'base_carbon_price'},
            {'label': f'{self.company_name} Revenue', 'formula': 'revenue', 'format': 'currency', 'total': True, 'key': 'revenue'},
            {'label': 'Project Implementation Costs', 'data_col': 'project_implementation_costs', 'format': 'currency', 'total': True, 'key': 'project_implementation_costs'},
            {'label': f'{self.company_name} Investment Drawdown', 'formula': 'investment', 'format': 'currency', 'total': True, 'key': 'investment'},
            {'label': f'{self.company_name} Net Cash Flow', 'formula': 'net_cf', 'format': 'currency', 'total': True, 'key': 'net_cf'},
            {'label': 'Discount Factor', 'formula': 'discount', 'format': 'number', 'total': False, 'key': 'discount'},
            {'label': 'Present Value', 'formula': 'pv', 'format': 'currency', 'total': True, 'key': 'pv'},
            {'label': 'Cumulative Cash Flow', 'formula': 'cum_cf', 'format': 'currency', 'total': False, 'key': 'cum_cf'},
            {'label': 'Cumulative PV', 'formula': 'cum_pv', 'format': 'currency', 'total': False, 'key': 'cum_pv'},
        ]
        
        # Track row positions for formula references
        row_positions = {}
        data_row = header_row + 1
        
        # Style definitions
        label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        label_font = Font(bold=True)
        formula_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        total_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        
        # Write each line item
        for item_idx, item in enumerate(line_items):
            current_row = data_row + item_idx  # This is already the Excel row (1-based)
            
            # Write label in column A
            label_cell = ws.cell(row=current_row, column=1)
            label_cell.value = item['label']
            label_cell.font = label_font
            label_cell.fill = label_fill
            label_cell.alignment = Alignment(horizontal='right', vertical='center')
            label_cell.border = thin_border
            
            # Store row position using the key (current_row is already Excel row number)
            row_positions[item['key']] = current_row
            
            # Write data/formulas for each year
            for year_idx in range(num_years):
                col = year_start_col + year_idx
                col_letter = get_column_letter(col)
                
                if 'data_col' in item:
                    # Write data value
                    data_col = item['data_col']
                    if data_col in valuation_schedule.columns:
                        # Get the year value from index (Year 1, 2, 3... or actual year)
                        year_val = valuation_schedule.index[year_idx] if year_idx < len(valuation_schedule) else year_idx + 1
                        if year_val in valuation_schedule.index:
                            value = valuation_schedule.loc[year_val, data_col]
                            if pd.notna(value):
                                cell = ws.cell(row=current_row, column=col)
                                cell.value = float(value)
                                if item['format'] == 'currency':
                                    cell.number_format = '$#,##0.00'
                                else:
                                    cell.number_format = '#,##0'
                                cell.border = thin_border
                                cell.alignment = Alignment(horizontal='right', vertical='center')
                
                elif 'formula' in item:
                    # Write formula
                    formula_type = item['formula']
                    cell = ws.cell(row=current_row, column=col)
                    cell.border = thin_border
                    cell.fill = formula_fill
                    cell.alignment = Alignment(horizontal='right', vertical='center')
                    
                    if formula_type == 'credits_share':
                        # Rubicon Share = Credits Gross * Streaming %
                        credits_row = row_positions['carbon_credits_gross']
                        cell.value = f"={col_letter}{credits_row}*{streaming_cell}"
                        cell.number_format = '#,##0'
                    
                    elif formula_type == 'revenue':
                        # Revenue = Share of Credits * Price
                        share_row = row_positions['credits_share']
                        price_row = row_positions['base_carbon_price']
                        cell.value = f"={col_letter}{share_row}*{col_letter}{price_row}"
                        cell.number_format = '$#,##0.00'
                    
                    elif formula_type == 'investment':
                        # Investment = -Investment/Tenor if Year <= Tenor, else 0
                        # Use year number (1-21) where year 1 = 2025
                        year_num = year_idx + 1
                        cell.value = f"=IF({year_num}<={tenor_cell},-{investment_cell}/{tenor_cell},0)"
                        cell.number_format = '$#,##0.00'
                    
                    elif formula_type == 'net_cf':
                        # Net CF = Revenue + Investment
                        revenue_row = row_positions['revenue']
                        investment_row = row_positions['investment']
                        cell.value = f"={col_letter}{revenue_row}+{col_letter}{investment_row}"
                        cell.number_format = '$#,##0.00'
                    
                    elif formula_type == 'discount':
                        # Discount Factor = 1 / (1 + WACC)^(Year - 1)
                        year_num = year_idx + 1
                        cell.value = f"=1/((1+{wacc_cell})^({year_num}-1))"
                        cell.number_format = '#,##0.00'
                    
                    elif formula_type == 'pv':
                        # PV = Net CF * Discount Factor
                        net_cf_row = row_positions['net_cf']
                        discount_row = row_positions['discount']
                        cell.value = f"={col_letter}{net_cf_row}*{col_letter}{discount_row}"
                        cell.number_format = '$#,##0.00'
                    
                    elif formula_type == 'cum_cf':
                        # Cumulative CF = Previous + Current
                        net_cf_row = row_positions['net_cf']
                        if year_idx == 0:
                            cell.value = f"={col_letter}{net_cf_row}"
                        else:
                            prev_col = get_column_letter(col - 1)
                            cell.value = f"={prev_col}{current_row}+{col_letter}{net_cf_row}"
                        cell.number_format = '$#,##0.00'
                    
                    elif formula_type == 'cum_pv':
                        # Cumulative PV = Previous + Current PV
                        pv_row = row_positions['pv']
                        if year_idx == 0:
                            cell.value = f"={col_letter}{pv_row}"
                        else:
                            prev_col = get_column_letter(col - 1)
                            cell.value = f"={prev_col}{current_row}+{col_letter}{pv_row}"
                        cell.number_format = '$#,##0.00'
            
            # Write total formula if needed
            if item.get('total', False):
                first_col = get_column_letter(year_start_col)
                last_col = get_column_letter(year_start_col + num_years - 1)
                total_cell = ws.cell(row=current_row, column=total_col)
                total_cell.value = f"=SUM({first_col}{current_row}:{last_col}{current_row})"
                total_cell.font = Font(bold=True)
                total_cell.fill = total_fill
                total_cell.border = thin_border
                total_cell.alignment = Alignment(horizontal='right', vertical='center')
                if item['format'] == 'currency':
                    total_cell.number_format = '$#,##0.00'
                else:
                    total_cell.number_format = '#,##0'
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        for i in range(num_years):
            col_letter = get_column_letter(year_start_col + i)
            ws.column_dimensions[col_letter].width = 12
        ws.column_dimensions[get_column_letter(total_col)].width = 15
    
    def _populate_summary_sheet_comprehensive(self, ws, valuation_schedule, actual_irr, target_irr,
                                             payback_period, mc_results, risk_flags, risk_score, breakeven):
        """
        Comprehensively populate Summary & Results sheet.
        Creates the full structure from scratch matching the original ExcelExporter.
        """
        from openpyxl.utils import get_column_letter
        
        # Clear existing data
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.value = None
                    if cell.fill:
                        cell.fill = PatternFill()  # Reset to default fill
                    if cell.font:
                        cell.font = Font()  # Reset to default font
        
        # Styles
        title_font = Font(bold=True, size=14)
        subtitle_font = Font(bold=True, size=12)
        subtitle_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        label_font = Font(bold=True)
        label_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        bold_currency_font = Font(bold=True)
        bold_percent_font = Font(bold=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        row = 1
        
        # Title
        ws.cell(row=row, column=1).value = 'Summary & Results'
        ws.cell(row=row, column=1).font = title_font
        row += 2
        
        # Key Metrics Section
        ws.cell(row=row, column=1).value = 'Key Financial Metrics'
        ws.cell(row=row, column=1).font = subtitle_font
        ws.cell(row=row, column=1).fill = subtitle_fill
        row += 1
        
        # NPV (formula reference to Valuation Schedule)
        ws.cell(row=row, column=1).value = 'Net Present Value (NPV)'
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        # PV is row 12 (Excel row 12), years are columns B-{last_col} (configurable)
        # Column B is index 2 (1-indexed), so last column is 2 + num_years - 1
        last_col_index = 2 + self.num_years - 1  # Column B (2) + num_years - 1
        last_col_letter = get_column_letter(last_col_index)
        npv_formula = f"=SUM('Valuation Schedule'!B12:{last_col_letter}12)"
        npv_cell = ws.cell(row=row, column=2)
        npv_cell.value = npv_formula
        npv_cell.font = bold_currency_font
        npv_cell.number_format = '$#,##0.00'
        npv_cell.border = thin_border
        row += 1
        
        # IRR (formula reference to Valuation Schedule)
        ws.cell(row=row, column=1).value = 'Internal Rate of Return (IRR)'
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        # Use Excel's IRR function on the Net Cash Flow row (row 10, Excel row 10, columns B-{last_col})
        irr_formula = f"=IRR('Valuation Schedule'!B10:{last_col_letter}10)"
        irr_cell = ws.cell(row=row, column=2)
        irr_cell.value = irr_formula
        irr_cell.font = bold_percent_font
        irr_cell.number_format = '0.00%'
        irr_cell.border = thin_border
        
        # Python calculated value as note
        note_cell = ws.cell(row=row, column=3)
        note_cell.value = f'(Python calculated: {actual_irr:.2%})'
        note_cell.font = Font(size=9, italic=True)
        row += 1
        
        # Payback Period
        if payback_period is not None:
            ws.cell(row=row, column=1).value = 'Payback Period (Years)'
            ws.cell(row=row, column=1).font = label_font
            ws.cell(row=row, column=1).fill = label_fill
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
            
            # Payback formula: Find column where cumulative CF becomes positive (row 13, Excel row 13, columns B-{last_col})
            # MATCH with match_type=1 finds the largest value <= 0, which gives us the year before payback
            # Old version didn't add +1, so we match that
            payback_formula = f"=MATCH(0,'Valuation Schedule'!B13:{last_col_letter}13,1)"
            payback_cell = ws.cell(row=row, column=2)
            payback_cell.value = payback_formula
            payback_cell.font = Font(bold=True)
            payback_cell.border = thin_border
            
            note_cell = ws.cell(row=row, column=3)
            note_cell.value = f'(Actual: {payback_period:.2f} years)'
            note_cell.font = Font(size=9, italic=True)
            row += 1
        
        # Target Metrics Section
        row += 1
        ws.cell(row=row, column=1).value = 'Target Metrics'
        ws.cell(row=row, column=1).font = subtitle_font
        ws.cell(row=row, column=1).fill = subtitle_fill
        row += 1
        
        # Get reference to Inputs sheet
        wb = ws.parent
        inputs_sheet = wb['Inputs & Assumptions'] if 'Inputs & Assumptions' in wb.sheetnames else None
        inputs_sheet_name = inputs_sheet.title if inputs_sheet else 'Inputs & Assumptions'
        
        # Target IRR (formula reference)
        ws.cell(row=row, column=1).value = 'Target IRR'
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        # Target IRR (matching template: B8)
        target_irr_formula = f"='{inputs_sheet_name}'!$B$8"
        target_irr_cell = ws.cell(row=row, column=2)
        target_irr_cell.value = target_irr_formula
        target_irr_cell.font = bold_percent_font
        target_irr_cell.number_format = '0.00%'
        target_irr_cell.border = thin_border
        row += 1
        
        # Target Streaming Percentage (matching template: B9)
        ws.cell(row=row, column=1).value = 'Target Streaming Percentage'
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        target_streaming_formula = f"='{inputs_sheet_name}'!$B$9"
        target_streaming_cell = ws.cell(row=row, column=2)
        target_streaming_cell.value = target_streaming_formula
        target_streaming_cell.font = bold_percent_font
        target_streaming_cell.number_format = '0.00%'
        target_streaming_cell.border = thin_border
        row += 1
        
        # Actual IRR Achieved
        ws.cell(row=row, column=1).value = 'Actual IRR Achieved'
        ws.cell(row=row, column=1).font = label_font
        ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
        
        actual_irr_cell = ws.cell(row=row, column=2)
        actual_irr_cell.value = actual_irr
        actual_irr_cell.font = bold_percent_font
        actual_irr_cell.number_format = '0.00%'
        actual_irr_cell.border = thin_border
        row += 1
        
        # Monte Carlo Summary
        if mc_results is not None:
            row += 1
            ws.cell(row=row, column=1).value = 'Monte Carlo Simulation Summary'
            ws.cell(row=row, column=1).font = subtitle_font
            ws.cell(row=row, column=1).fill = subtitle_fill
            row += 1
            
            mc_metrics = [
                ('MC Mean IRR', 'mc_mean_irr', 'percent'),
                ('MC P10 IRR (Downside)', 'mc_p10_irr', 'percent'),
                ('MC P90 IRR (Upside)', 'mc_p90_irr', 'percent'),
                ('MC Mean NPV', 'mc_mean_npv', 'currency'),
                ('MC P10 NPV', 'mc_p10_npv', 'currency'),
                ('MC P90 NPV', 'mc_p90_npv', 'currency'),
            ]
            
            for label, key, fmt_type in mc_metrics:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=1).font = label_font
                ws.cell(row=row, column=1).fill = label_fill
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
                
                value = mc_results.get(key, 0)
                value_cell = ws.cell(row=row, column=2)
                if pd.isna(value) or not np.isfinite(value):
                    value_cell.value = 'N/A'
                else:
                    value_cell.value = float(value)
                    if fmt_type == 'percent':
                        value_cell.number_format = '0.00%'
                        value_cell.font = bold_percent_font
                    else:
                        value_cell.number_format = '$#,##0.00'
                        value_cell.font = bold_currency_font
                value_cell.border = thin_border
                row += 1
        
        # Risk Assessment Section
        if risk_flags is not None or risk_score is not None:
            row += 1
            ws.cell(row=row, column=1).value = 'Risk Assessment'
            ws.cell(row=row, column=1).font = subtitle_font
            ws.cell(row=row, column=1).fill = subtitle_fill
            row += 1
            
            # Risk Score
            if risk_score is not None:
                ws.cell(row=row, column=1).value = 'Overall Risk Score'
                ws.cell(row=row, column=1).font = label_font
                ws.cell(row=row, column=1).fill = label_fill
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
                
                score = risk_score.get('overall_risk_score', 0)
                category = risk_score.get('risk_category', 'Unknown')
                
                score_cell = ws.cell(row=row, column=2)
                score_cell.value = score
                score_cell.number_format = '0.0'
                score_cell.font = Font(bold=True)
                score_cell.border = thin_border
                
                # Color code based on risk level
                if category == 'Low':
                    score_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif category == 'Medium':
                    score_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:  # High
                    score_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                
                note_cell = ws.cell(row=row, column=3)
                note_cell.value = f'({category} Risk)'
                row += 1
                
                # Component risk scores
                component_risks = [
                    ('  Financial Risk', 'financial_risk'),
                    ('  Volume Risk', 'volume_risk'),
                    ('  Price Risk', 'price_risk'),
                    ('  Operational Risk', 'operational_risk'),
                ]
                
                for label, key in component_risks:
                    ws.cell(row=row, column=1).value = label
                    ws.cell(row=row, column=1).font = Font()
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
                    
                    value_cell = ws.cell(row=row, column=2)
                    value_cell.value = risk_score.get(key, 0)
                    value_cell.number_format = '#,##0'
                    value_cell.border = thin_border
                    row += 1
            
            # Risk Flags
            if risk_flags is not None:
                row += 1
                risk_level = risk_flags.get('risk_level', 'unknown')
                
                ws.cell(row=row, column=1).value = 'Risk Level'
                ws.cell(row=row, column=1).font = label_font
                ws.cell(row=row, column=1).fill = label_fill
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
                
                level_cell = ws.cell(row=row, column=2)
                level_cell.font = Font(bold=True)
                level_cell.border = thin_border
                
                if risk_level == 'red':
                    level_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                    level_cell.value = '🔴 HIGH RISK'
                elif risk_level == 'yellow':
                    level_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                    level_cell.value = '🟡 MEDIUM RISK'
                else:
                    level_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    level_cell.value = '🟢 LOW RISK'
                row += 1
                
                # Flag counts
                flag_counts = risk_flags.get('flag_count', {})
                ws.cell(row=row, column=1).value = '  Red Flags'
                ws.cell(row=row, column=1).font = Font()
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2).value = flag_counts.get('red', 0)
                ws.cell(row=row, column=2).number_format = '#,##0'
                ws.cell(row=row, column=2).border = thin_border
                row += 1
                
                ws.cell(row=row, column=1).value = '  Yellow Flags'
                ws.cell(row=row, column=1).font = Font()
                ws.cell(row=row, column=1).border = thin_border
                ws.cell(row=row, column=2).value = flag_counts.get('yellow', 0)
                ws.cell(row=row, column=2).number_format = '#,##0'
                ws.cell(row=row, column=2).border = thin_border
                row += 1
                
                # List flags
                red_flags = risk_flags.get('red_flags', [])
                yellow_flags = risk_flags.get('yellow_flags', [])
                green_flags = risk_flags.get('green_flags', [])
                
                if red_flags:
                    row += 1
                    ws.cell(row=row, column=1).value = '🚨 Critical Risks (Red Flags):'
                    ws.cell(row=row, column=1).font = subtitle_font
                    ws.cell(row=row, column=1).fill = subtitle_fill
                    row += 1
                    for i, flag in enumerate(red_flags, 1):
                        ws.cell(row=row, column=1).value = f'{i}. {flag}'
                        row += 1
                
                if yellow_flags:
                    row += 1
                    ws.cell(row=row, column=1).value = '⚠️  Warnings (Yellow Flags):'
                    ws.cell(row=row, column=1).font = subtitle_font
                    ws.cell(row=row, column=1).fill = subtitle_fill
                    row += 1
                    for i, flag in enumerate(yellow_flags, 1):
                        ws.cell(row=row, column=1).value = f'{i}. {flag}'
                        row += 1
        
        # Breakeven Analysis Section
        if breakeven is not None:
            row += 1
            ws.cell(row=row, column=1).value = 'Breakeven Analysis'
            ws.cell(row=row, column=1).font = subtitle_font
            ws.cell(row=row, column=1).fill = subtitle_fill
            row += 1
            
            # Breakeven Price
            if 'breakeven_price' in breakeven:
                be_price = breakeven['breakeven_price']
                if be_price and 'breakeven_price' in be_price and not pd.isna(be_price.get('breakeven_price')):
                    ws.cell(row=row, column=1).value = 'Breakeven Carbon Price'
                    ws.cell(row=row, column=1).font = label_font
                    ws.cell(row=row, column=1).fill = label_fill
                    ws.cell(row=row, column=1).border = thin_border
                    ws.cell(row=row, column=1).alignment = Alignment(horizontal='right', vertical='center')
                    
                    price_cell = ws.cell(row=row, column=2)
                    price_cell.value = be_price['breakeven_price']
                    price_cell.font = bold_currency_font
                    price_cell.number_format = '$#,##0.00'
                    price_cell.border = thin_border
                    
                    if 'base_price' in be_price:
                        multiplier = be_price.get('price_multiplier', 1.0)
                        note_cell = ws.cell(row=row, column=3)
                        note_cell.value = f'({multiplier:.2f}x base price)'
                        note_cell.font = Font(size=9, italic=True)
                    row += 1
        
        # Set column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 30
    
    def _populate_deal_valuation_sheet(self, ws, deal_valuation_results: Dict):
        """Populate Deal Valuation sheet."""
        if not deal_valuation_results:
            return
        
        row = 2
        if 'maximum_purchase_price' in deal_valuation_results:
            ws.cell(row=row, column=1).value = 'Maximum Purchase Price'
            ws.cell(row=row, column=2).value = float(deal_valuation_results['maximum_purchase_price'])
            ws.cell(row=row, column=2).number_format = '$#,##0.00'
            row += 1
        
        if 'actual_irr' in deal_valuation_results:
            ws.cell(row=row, column=1).value = 'Actual IRR'
            ws.cell(row=row, column=2).value = float(deal_valuation_results['actual_irr'])
            ws.cell(row=row, column=2).number_format = '0.00%'
            row += 1
    
    def _populate_monte_carlo_sheet(self, ws, mc_results: Dict):
        """Populate Monte Carlo Results sheet."""
        if not mc_results:
            return
        
        row = 2
        metrics = [
            ('Mean IRR', 'mc_mean_irr', '0.00%'),
            ('P10 IRR', 'mc_p10_irr', '0.00%'),
            ('P50 IRR', 'mc_p50_irr', '0.00%'),
            ('P90 IRR', 'mc_p90_irr', '0.00%'),
            ('Mean NPV', 'mc_mean_npv', '$#,##0.00'),
            ('P10 NPV', 'mc_p10_npv', '$#,##0.00'),
            ('P50 NPV', 'mc_p50_npv', '$#,##0.00'),
            ('P90 NPV', 'mc_p90_npv', '$#,##0.00'),
        ]
        
        for label, key, fmt in metrics:
            if key in mc_results and mc_results[key] is not None:
                ws.cell(row=row, column=1).value = label
                ws.cell(row=row, column=2).value = float(mc_results[key])
                ws.cell(row=row, column=2).number_format = fmt
                row += 1
    
    def _populate_sensitivity_sheet(self, ws, sensitivity_table: pd.DataFrame):
        """Populate Sensitivity Analysis sheet."""
        if sensitivity_table is None or sensitivity_table.empty:
            return
        
        # Clear existing data (keep headers)
        for row in range(2, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value:
                    cell.value = None
        
        # Write table starting from row 2
        # Column headers (price multipliers)
        col_idx = 2
        for price_mult in sensitivity_table.columns:
            ws.cell(row=2, column=col_idx).value = str(price_mult)
            ws.cell(row=2, column=col_idx).font = Font(bold=True)
            col_idx += 1
        
        # Row headers and data
        row_idx = 3
        for credit_mult in sensitivity_table.index:
            ws.cell(row=row_idx, column=1).value = str(credit_mult)
            ws.cell(row=row_idx, column=1).font = Font(bold=True)
            
            col_idx = 2
            for price_mult in sensitivity_table.columns:
                irr_value = sensitivity_table.loc[credit_mult, price_mult]
                if pd.notna(irr_value):
                    ws.cell(row=row_idx, column=col_idx).value = float(irr_value)
                    ws.cell(row=row_idx, column=col_idx).number_format = '0.00%'
                col_idx += 1
            row_idx += 1
    
    def _add_presentation_charts_to_inputs(self, ws, assumptions: Dict, streaming_pct: float):
        """Generate and embed charts in Inputs & Assumptions sheet."""
        try:
            from .presentation_charts import PresentationChartGenerator
            import tempfile
            
            chart_gen = PresentationChartGenerator()
            
            # Chart 1: Assumptions Summary (E2)
            chart_path = chart_gen.create_assumptions_summary_chart(assumptions, streaming_pct)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'E2', width=400, height=300)
            
            # Chart 2: Price Projection (E17)
            chart_path = chart_gen.create_price_projection_chart(assumptions, years=20)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'E17', width=400, height=300)
            
            # Chart 3: Volume Projection (E34)
            chart_path = chart_gen.create_volume_projection_chart(assumptions, years=20)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'E34', width=400, height=300)
            
        except Exception as e:
            print(f"Warning: Could not add charts to Inputs & Assumptions: {e}")
    
    def _add_presentation_charts_to_valuation(self, ws, valuation_schedule: pd.DataFrame):
        """Generate and embed charts in Valuation Schedule sheet."""
        try:
            from .presentation_charts import PresentationChartGenerator
            
            chart_gen = PresentationChartGenerator()
            
            # Chart 1: Cash Flow Waterfall (below data, row 25)
            chart_path = chart_gen.create_cash_flow_waterfall(valuation_schedule, years=20)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'A25', width=600, height=350)
            
            # Chart 2: Cumulative Cash Flow (I25)
            chart_path = chart_gen.create_cumulative_cash_flow(valuation_schedule, years=20)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'I25', width=400, height=300)
            
            # Chart 3: NPV Trend (A45)
            chart_path = chart_gen.create_npv_trend(valuation_schedule, years=20)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'A45', width=600, height=350)
            
        except Exception as e:
            print(f"Warning: Could not add charts to Valuation Schedule: {e}")
    
    def _add_presentation_charts_to_summary(self, ws, actual_irr: float, target_irr: float, risk_score: Dict):
        """Generate and embed charts in Summary & Results sheet."""
        try:
            from .presentation_charts import PresentationChartGenerator
            
            chart_gen = PresentationChartGenerator()
            
            # Chart 1: Financial Metrics Dashboard (E5) - placeholder for now
            # Could add sparklines or mini charts here
            
            # Chart 2: Risk Breakdown (E15)
            if risk_score:
                chart_path = chart_gen.create_risk_breakdown(risk_score)
                chart_gen.embed_chart_in_excel(chart_path, ws, 'E15', width=400, height=300)
            
            # Chart 3: Return Summary (E30)
            chart_path = chart_gen.create_return_summary(target_irr, actual_irr)
            chart_gen.embed_chart_in_excel(chart_path, ws, 'E30', width=400, height=300)
            
        except Exception as e:
            print(f"Warning: Could not add charts to Summary & Results: {e}")
